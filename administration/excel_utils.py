from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from administration.excel_schema import (
    STUDENT_EXCEL_COLUMNS,
    FACULTY_EXCEL_COLUMNS,
    STUDENT_TEMPLATE_METADATA_ROWS,
    STUDENT_TEMPLATE_VERSION,
    FACULTY_TEMPLATE_METADATA_ROWS,
    FACULTY_TEMPLATE_VERSION,
)

# =========================================================
# INTERNAL BASE TEMPLATE (VERSION SAFE)
# =========================================================
def _generate_template(file_path, sheet_name, columns):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="DDDDDD")

    # ----------------------------
    # HEADERS (LOCKED)
    # ----------------------------
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.protection = cell.protection.copy(locked=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # ----------------------------
    # DATA ROWS (UNLOCKED)
    # ----------------------------
    for row in range(2, 2000):
        for col in range(1, len(columns) + 1):
            data_cell = ws.cell(row=row, column=col)
            data_cell.protection = data_cell.protection.copy(locked=False)

    # ----------------------------
    # ENABLE SHEET PROTECTION (SAFE)
    # ----------------------------
    ws.protection.enable()
    ws.protection.set_password("college-erp")

    # ❗ DO NOT set unsupported flags
    # Excel by default allows editing unlocked cells

    wb.save(file_path)


def _student_metadata_defaults(metadata=None):
    incoming = metadata or {}
    return {
        "College Name": str(incoming.get("College Name", "") or "").strip(),
        "Branch": str(incoming.get("Branch", "") or "").strip(),
        "Academic Year": str(incoming.get("Academic Year", "") or "").strip(),
        "Batch": str(
            incoming.get("Batch", incoming.get("Academic Year", "")) or ""
        ).strip(),
        "Semester": str(incoming.get("Semester", "") or "").strip(),
        "Template Version": str(
            incoming.get("Template Version", STUDENT_TEMPLATE_VERSION)
            or STUDENT_TEMPLATE_VERSION
        ).strip(),
    }


def _write_student_sheet_with_metadata(file_path, metadata, data_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    meta_font = Font(bold=True)
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="FFF59D")
    meta_fill = PatternFill(fill_type="solid", fgColor="FFF59D")

    safe_metadata = _student_metadata_defaults(metadata)

    # Rows 1-5 => metadata (Label in A, Value in B)
    for idx, label in enumerate(STUDENT_TEMPLATE_METADATA_ROWS, start=1):
        label_cell = ws.cell(row=idx, column=1, value=label)
        label_cell.font = meta_font
        label_cell.fill = meta_fill
        ws.cell(row=idx, column=2, value=safe_metadata.get(label, ""))

    # Header row starts after metadata rows
    header_row = len(STUDENT_TEMPLATE_METADATA_ROWS) + 1
    section_fills = {
        "ADMISSION": PatternFill(fill_type="solid", fgColor="FFF59D"),
        "PERSONAL": PatternFill(fill_type="solid", fgColor="FFE082"),
        "IDENTIFICATION": PatternFill(fill_type="solid", fgColor="FFD54F"),
        "CET": PatternFill(fill_type="solid", fgColor="FFEE58"),
        "FAMILY_CONTACT": PatternFill(fill_type="solid", fgColor="FFF176"),
        "ADDRESS": PatternFill(fill_type="solid", fgColor="FFECB3"),
    }
    column_sections = {
        "S.No": "ADMISSION",
        "Roll No": "ADMISSION",
        "Admission No": "ADMISSION",
        "Student Name as per SSC": "PERSONAL",
        "Gender": "PERSONAL",
        "Date of Birth": "PERSONAL",
        "Admission Date": "ADMISSION",
        "Admission Type": "ADMISSION",
        "Batch": "ADMISSION",
        "Identification Mark 1": "IDENTIFICATION",
        "Identification Mark 2": "IDENTIFICATION",
        "CET Type": "CET",
        "CET HT No": "CET",
        "CET Rank": "CET",
        "Lateral Entry": "CET",
        "Branch": "ADMISSION",
        "Completion Year": "ADMISSION",
        "Father Name": "FAMILY_CONTACT",
        "Mother Name": "FAMILY_CONTACT",
        "Father Aadhaar No": "IDENTIFICATION",
        "Mother Aadhaar No": "IDENTIFICATION",
        "Student Aadhaar No": "IDENTIFICATION",
        "Student Mobile No": "FAMILY_CONTACT",
        "Father Mobile No": "FAMILY_CONTACT",
        "Religion": "PERSONAL",
        "Caste": "PERSONAL",
        "Sub-Caste": "PERSONAL",
        "Scholarship": "ADMISSION",
        "Fee Reimbursement Amount": "ADMISSION",
        "Do.No 1": "ADDRESS",
        "Village 1": "ADDRESS",
        "Mandal 1": "ADDRESS",
        "District 1": "ADDRESS",
        "Pincode 1": "ADDRESS",
        "Do.No 2": "ADDRESS",
        "Village 2": "ADDRESS",
        "Mandal 2": "ADDRESS",
        "District 2": "ADDRESS",
        "Pincode 2": "ADDRESS",
    }
    for col_idx, col_name in enumerate(STUDENT_EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font
        section = column_sections.get(col_name, "ADMISSION")
        cell.fill = section_fills.get(section, header_fill)
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # Data starts after header row
    data_start_row = header_row + 1
    for row_offset, row_data in enumerate(data_rows):
        excel_row = data_start_row + row_offset
        for col_idx, col_name in enumerate(STUDENT_EXCEL_COLUMNS, start=1):
            ws.cell(row=excel_row, column=col_idx, value=row_data.get(col_name, ""))

    # Dropdown validations
    dropdowns = [
        ("Gender", ["Male", "Female", "Other"]),
        ("CET Type", ["EAPCET", "ECET"]),
        ("Lateral Entry", ["Yes", "No"]),
        ("Admission Type", ["Convenor", "Management"]),
    ]
    col_idx_map = {name: idx for idx, name in enumerate(STUDENT_EXCEL_COLUMNS, start=1)}
    for col_name, values in dropdowns:
        col_idx = col_idx_map[col_name]
        col_letter = get_column_letter(col_idx)
        list_formula = '"' + ",".join(values) + '"'
        dv = DataValidation(type="list", formula1=list_formula, allow_blank=True)
        dv.promptTitle = f"{col_name} Selection"
        dv.prompt = f"Choose one value for {col_name}"
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}{data_start_row}:{col_letter}2000")

    # Length validations
    for col_name, length in [
        ("Student Mobile No", 10),
        ("Father Mobile No", 10),
        ("Student Aadhaar No", 12),
        ("Father Aadhaar No", 12),
        ("Mother Aadhaar No", 12),
    ]:
        col_idx = col_idx_map[col_name]
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(
            type="textLength",
            operator="equal",
            formula1=str(length),
            allow_blank=True,
        )
        dv.errorTitle = "Invalid Length"
        dv.error = f"{col_name} must be exactly {length} digits"
        dv.promptTitle = f"{col_name} Validation"
        dv.prompt = f"Enter exactly {length} digits"
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}{data_start_row}:{col_letter}2000")

    wb.save(file_path)


# =========================================================
# STUDENT EXCEL TEMPLATE
# =========================================================
def generate_student_template(file_path, metadata=None):
    _write_student_sheet_with_metadata(
        file_path=file_path,
        metadata=metadata,
        data_rows=[],
    )


def generate_student_export(file_path, metadata=None, data_rows=None):
    _write_student_sheet_with_metadata(
        file_path=file_path,
        metadata=metadata,
        data_rows=data_rows or [],
    )


# =========================================================
# FACULTY EXCEL TEMPLATE
# =========================================================
def _faculty_metadata_defaults(metadata=None):
    incoming = metadata or {}
    return {
        "College Name": str(incoming.get("College Name", "SVR Engineering College") or "SVR Engineering College").strip(),
        "Department": str(incoming.get("Department", "") or "").strip(),
        "Designation": str(incoming.get("Designation", "") or "").strip(),
        "Template Version": str(
            incoming.get("Template Version", FACULTY_TEMPLATE_VERSION)
            or FACULTY_TEMPLATE_VERSION
        ).strip(),
    }


def _write_faculty_sheet_with_metadata(file_path, metadata):
    wb = Workbook()
    ws = wb.active
    ws.title = "Faculty"

    meta_font = Font(bold=True)
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="FFF59D")
    meta_fill = PatternFill(fill_type="solid", fgColor="FFF59D")

    safe_metadata = _faculty_metadata_defaults(metadata)

    # Rows 1-4 => metadata (Label in A, Value in B)
    for idx, label in enumerate(FACULTY_TEMPLATE_METADATA_ROWS, start=1):
        label_cell = ws.cell(row=idx, column=1, value=label)
        label_cell.font = meta_font
        label_cell.fill = meta_fill
        label_cell.protection = label_cell.protection.copy(locked=True)

        value_cell = ws.cell(row=idx, column=2, value=safe_metadata.get(label, ""))
        value_cell.fill = meta_fill
        value_cell.protection = value_cell.protection.copy(locked=True)

    # Header row starts after metadata rows
    header_row = len(FACULTY_TEMPLATE_METADATA_ROWS) + 1
    for col_idx, col_name in enumerate(FACULTY_EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.protection = cell.protection.copy(locked=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # Data rows unlocked
    data_start_row = header_row + 1
    for row in range(data_start_row, 2000):
        for col in range(1, len(FACULTY_EXCEL_COLUMNS) + 1):
            data_cell = ws.cell(row=row, column=col)
            data_cell.protection = data_cell.protection.copy(locked=False)

    # Gender dropdown
    col_idx_map = {name: idx for idx, name in enumerate(FACULTY_EXCEL_COLUMNS, start=1)}
    gender_col = get_column_letter(col_idx_map["Gender"])
    dv = DataValidation(type="list", formula1='"MALE,FEMALE,OTHER"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{gender_col}{data_start_row}:{gender_col}2000")

    # Protect sheet so metadata/header remain uneditable
    ws.protection.enable()
    ws.protection.set_password("college-erp")

    wb.save(file_path)


def generate_faculty_template(file_path, metadata=None):
    _write_faculty_sheet_with_metadata(file_path=file_path, metadata=metadata or {})
