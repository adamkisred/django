import os
import tempfile
import pandas as pd
from datetime import datetime
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook

from django.db import IntegrityError, transaction
from django.db.models import Q
from django.http import FileResponse
from django.contrib.auth import authenticate
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.utils import timezone

from rest_framework.views import APIView
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework import status

from accounts.models import User
from students.models import Student
from faculty.models import Faculty

from administration.excel_schema import (
    STUDENT_EXCEL_COLUMNS,
    FACULTY_EXCEL_COLUMNS,
    STUDENT_TEMPLATE_METADATA_ROWS,
    STUDENT_TEMPLATE_VERSION,
    FACULTY_TEMPLATE_METADATA_ROWS,
    FACULTY_TEMPLATE_VERSION,
)
from administration.excel_utils import (
    generate_student_template,
    generate_student_export,
    generate_faculty_template,
)

# =========================================================
# COMMON DATE PARSER
# =========================================================
def parse_date(value):
    if value is None or pd.isna(value) or str(value).strip() == "":
        return None

    if isinstance(value, (datetime, pd.Timestamp)):
        return value.date()

    value = str(value).strip()

    for fmt in ("%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue

    try:
        return pd.to_datetime(value, unit="D", origin="1899-12-30").date()
    except Exception:
        return None


def parse_decimal(value):
    if value is None or pd.isna(value):
        return None
    raw = str(value).strip()
    if raw == "":
        return None
    try:
        return Decimal(raw)
    except (InvalidOperation, ValueError):
        return None


def _normalize_faculty_id(value):
    if value is None or pd.isna(value):
        return ""
    return str(value).strip().upper()


def _clean_text(value):
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def _build_address(*parts):
    return " | ".join([p for p in (_clean_text(x) for x in parts) if p])


def _resolve_admin_user(request):
    current = request.user if getattr(request, "user", None) and request.user.is_authenticated else None
    if current and getattr(current, "role", None) == "ADMIN":
        return current

    admin_identifier = (request.data.get("admin_identifier") or "").strip()
    admin_password = request.data.get("admin_password") or ""
    if not admin_identifier or not admin_password:
        return None

    admin_user = authenticate(request, username=admin_identifier, password=admin_password)
    if not admin_user:
        return None

    if getattr(admin_user, "role", None) != "ADMIN":
        return None

    return admin_user


def _clean_meta(value):
    return str(value or "").strip()


def _normalized_compact(value):
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def _matches_compact(value, selected, contains=False):
    selected_norm = _normalized_compact(selected)
    if not selected_norm:
        return True
    value_norm = _normalized_compact(value)
    if contains:
        return selected_norm in value_norm
    return value_norm == selected_norm


def _selected_student_context(request):
    academic_year = _clean_meta(
        request.data.get("academic_year")
        or request.query_params.get("academic_year")
    )
    return {
        "college_name": _clean_meta(
            request.data.get("college_name")
            or request.query_params.get("college_name")
            or "SVREC"
        ),
        "branch": _clean_meta(
            request.data.get("branch")
            or request.query_params.get("branch")
        ),
        "academic_year": academic_year,
        "batch": _clean_meta(
            request.data.get("batch")
            or request.query_params.get("batch")
            or academic_year
        ),
        "semester": _clean_meta(
            request.data.get("semester")
            or request.query_params.get("semester")
        ),
    }


def _validate_student_context_fields(context):
    missing = []
    if not context["branch"]:
        missing.append("branch")
    if not context["academic_year"]:
        missing.append("academic_year")
    if not context["semester"]:
        missing.append("semester")
    return missing


def _read_student_sheet_metadata(file_obj):
    file_obj.seek(0)
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    metadata = {
        "college_name": _clean_meta(ws.cell(row=1, column=2).value),
        "branch": _clean_meta(ws.cell(row=2, column=2).value),
        "academic_year": _clean_meta(ws.cell(row=3, column=2).value),
        "batch": _clean_meta(ws.cell(row=4, column=2).value),
        "semester": _clean_meta(ws.cell(row=5, column=2).value),
        "template_version": _clean_meta(ws.cell(row=6, column=2).value),
    }
    wb.close()
    return metadata


def _selected_faculty_context(request):
    return {
        "college_name": _clean_meta(
            request.data.get("college_name")
            or request.query_params.get("college_name")
            or "SVR Engineering College"
        ),
        "department": _clean_meta(
            request.data.get("department")
            or request.query_params.get("department")
        ),
        "designation": _clean_meta(
            request.data.get("designation")
            or request.query_params.get("designation")
        ),
    }


def _validate_faculty_context_fields(context):
    missing = []
    if not context["department"]:
        missing.append("department")
    if not context["designation"]:
        missing.append("designation")
    return missing


def _read_faculty_sheet_metadata(file_obj):
    file_obj.seek(0)
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    metadata = {
        "college_name": _clean_meta(ws.cell(row=1, column=2).value),
        "department": _clean_meta(ws.cell(row=2, column=2).value),
        "designation": _clean_meta(ws.cell(row=3, column=2).value),
        "template_version": _clean_meta(ws.cell(row=4, column=2).value),
    }
    wb.close()
    return metadata


# =========================================================
# STUDENT TEMPLATE DOWNLOAD
# =========================================================
class StudentExcelTemplateDownload(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        context = _selected_student_context(request)
        missing = _validate_student_context_fields(context)
        if missing:
            return Response(
                {
                    "error": (
                        "Required query params missing: "
                        + ", ".join(missing)
                    )
                },
                status=400,
            )

        path = os.path.join(tempfile.gettempdir(), "student_upload_template.xlsx")
        generate_student_template(
            path,
            metadata={
                "College Name": context["college_name"],
                "Branch": context["branch"],
                "Academic Year": context["academic_year"],
                "Batch": context["batch"],
                "Semester": context["semester"],
                "Template Version": STUDENT_TEMPLATE_VERSION,
            },
        )
        return FileResponse(
            open(path, "rb"),
            as_attachment=True,
            filename="student_upload_template.xlsx",
        )


# =========================================================
# STUDENT EXCEL UPLOAD (EMAIL + OTP ONLY)
# =========================================================
class StudentExcelUpload(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        context = _selected_student_context(request)
        missing = _validate_student_context_fields(context)
        if missing:
            return Response(
                {
                    "error": (
                        "Please select and send: "
                        + ", ".join(missing)
                    )
                },
                status=400,
            )

        file = request.FILES.get("file")
        if not file:
            return Response({"error": "Excel file required"}, status=400)

        try:
            metadata = _read_student_sheet_metadata(file)
        except Exception:
            return Response(
                {"error": "Invalid student template. Could not read metadata rows."},
                status=400,
            )

        if metadata.get("template_version") != STUDENT_TEMPLATE_VERSION:
            return Response(
                {
                    "error": (
                        "Template version mismatch. "
                        "Please download a fresh template and upload again."
                    )
                },
                status=400,
            )

        mismatches = []
        if metadata.get("college_name", "").lower() != context["college_name"].lower():
            mismatches.append("college_name")
        if metadata.get("branch", "").lower() != context["branch"].lower():
            mismatches.append("branch")
        if metadata.get("academic_year", "").lower() != context["academic_year"].lower():
            mismatches.append("academic_year")
        if metadata.get("batch", "").lower() != context["batch"].lower():
            mismatches.append("batch")
        if metadata.get("semester", "").lower() != context["semester"].lower():
            mismatches.append("semester")

        if mismatches:
            return Response(
                {
                    "error": (
                        "Selected filters do not match template metadata: "
                        + ", ".join(mismatches)
                    ),
                    "selected": context,
                    "template_metadata": metadata,
                },
                status=400,
            )

        file.seek(0)
        df = pd.read_excel(file, header=len(STUDENT_TEMPLATE_METADATA_ROWS), dtype=str)
        df = df.fillna("")

        if list(df.columns) != STUDENT_EXCEL_COLUMNS:
            return Response(
                {
                    "error": (
                        "Invalid student data header row. "
                        "Do not change column names/order below metadata rows."
                    )
                },
                status=400,
            )

        created_count = 0
        updated_count = 0
        skipped_count = 0

        with transaction.atomic():
            for _, row in df.iterrows():
                roll_no = (row.get("Roll No") or "").strip().upper()
                if not roll_no:
                    skipped_count += 1
                    continue

                email = f"{roll_no.lower()}@svrec.ac.in"

                # 🔐 USER (STUDENT → EMAIL + OTP ONLY)
                user, created = User.objects.get_or_create(
                    email=email,
                    defaults={
                        "username": None,          # 🔑 IMPORTANT FIX
                        "role": "STUDENT",
                        "is_active": True,
                    },
                )

                if created:
                    user.set_unusable_password()
                    user.save(update_fields=["password"])

                _, student_created = Student.objects.update_or_create(
                    hall_ticket_no=roll_no,
                    defaults={
                        "user": user,
                        "email": email,
                        "admission_no": row.get("Admission No") or "",
                        "college_name": context["college_name"],
                        "branch": context["branch"],
                        "academic_year": context["academic_year"],
                        "semester": context["semester"],
                        "name": row.get("Student Name as per SSC") or "",
                        "gender": (row.get("Gender") or "")[:1].upper(),
                        "rank": int(row["CET Rank"])
                        if row.get("CET Rank") and str(row["CET Rank"]).isdigit()
                        else None,
                        "father_name": row.get("Father Name") or "",
                        "dob": parse_date(row.get("Date of Birth")),
                        "admission_date": parse_date(row.get("Admission Date")),
                        "student_mobile": row.get("Student Mobile No") or "",
                        "parent_mobile": row.get("Father Mobile No") or "",
                        "religion": row.get("Religion") or "",
                        "caste": row.get("Caste") or "",
                        "sub_caste": row.get("Sub-Caste") or "",
                        "aadhar": row.get("Student Aadhaar No") or "",
                        "address": "|".join([
                            "~".join([
                                str(row.get("Do.No 1") or "").strip(),
                                str(row.get("Village 1") or "").strip(),
                                str(row.get("Mandal 1") or "").strip(),
                                str(row.get("District 1") or "").strip(),
                                str(row.get("Pincode 1") or "").strip(),
                            ]),
                            "~".join([
                                str(row.get("Do.No 2") or "").strip(),
                                str(row.get("Village 2") or "").strip(),
                                str(row.get("Mandal 2") or "").strip(),
                                str(row.get("District 2") or "").strip(),
                                str(row.get("Pincode 2") or "").strip(),
                            ]),
                        ]).strip("|"),
                        "convenor": row.get("Admission Type") or "",
                    },
                )
                if student_created:
                    created_count += 1
                else:
                    updated_count += 1

        return Response(
            {
                "message": "Student Excel processed successfully",
                "summary": {
                    "created": created_count,
                    "updated": updated_count,
                    "skipped": skipped_count,
                },
                "context": context,
            },
            status=200,
        )


# =========================================================
# STUDENT EXCEL EXPORT
# =========================================================
class StudentExcelExport(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        context = _selected_student_context(request)
        missing = _validate_student_context_fields(context)
        if missing:
            return Response(
                {
                    "error": (
                        "Required query params missing: "
                        + ", ".join(missing)
                    )
                },
                status=400,
            )

        queryset = Student.objects.filter(
            branch=context["branch"],
            academic_year=context["academic_year"],
            semester=context["semester"],
        )
        if context["college_name"]:
            queryset = queryset.filter(college_name=context["college_name"])

        data = []
        for i, s in enumerate(queryset, start=1):
            do1 = village1 = mandal1 = district1 = pincode1 = ""
            do2 = village2 = mandal2 = district2 = pincode2 = ""
            if s.address:
                if "|" in s.address:
                    addr_blocks = s.address.split("|", 1)
                    block1 = addr_blocks[0]
                    block2 = addr_blocks[1] if len(addr_blocks) > 1 else ""
                    vals1 = (block1.split("~") + ["", "", "", "", ""])[:5]
                    vals2 = (block2.split("~") + ["", "", "", "", ""])[:5]
                    do1, village1, mandal1, district1, pincode1 = [v.strip() for v in vals1]
                    do2, village2, mandal2, district2, pincode2 = [v.strip() for v in vals2]
                else:
                    do1 = s.address.strip()
            data.append({
                "S.No": i,
                "Roll No": s.hall_ticket_no,
                "Admission No": s.admission_no,
                "Student Name as per SSC": s.name,
                "Gender": s.gender,
                "Date of Birth": s.dob.strftime("%d-%m-%Y") if s.dob else "",
                "Admission Date": s.admission_date.strftime("%d-%m-%Y") if s.admission_date else "",
                "Admission Type": s.convenor,
                "Batch": s.academic_year,
                "Identification Mark 1": "",
                "Identification Mark 2": "",
                "CET Type": "",
                "CET HT No": "",
                "CET Rank": s.rank,
                "Lateral Entry": "",
                "Branch": s.branch,
                "Completion Year": "",
                "Father Name": s.father_name,
                "Mother Name": "",
                "Father Aadhaar No": "",
                "Mother Aadhaar No": "",
                "Student Aadhaar No": s.aadhar,
                "Student Mobile No": s.student_mobile,
                "Father Mobile No": s.parent_mobile,
                "Religion": s.religion,
                "Caste": s.caste,
                "Sub-Caste": s.sub_caste,
                "Scholarship": "",
                "Fee Reimbursement Amount": "",
                "Do.No 1": do1,
                "Village 1": village1,
                "Mandal 1": mandal1,
                "District 1": district1,
                "Pincode 1": pincode1,
                "Do.No 2": do2,
                "Village 2": village2,
                "Mandal 2": mandal2,
                "District 2": district2,
                "Pincode 2": pincode2,
            })

        path = os.path.join(tempfile.gettempdir(), "students_export.xlsx")
        generate_student_export(
            path,
            metadata={
                "College Name": context["college_name"],
                "Branch": context["branch"],
                "Academic Year": context["academic_year"],
                "Batch": context["batch"],
                "Semester": context["semester"],
                "Template Version": STUDENT_TEMPLATE_VERSION,
            },
            data_rows=data,
        )

        return FileResponse(
            open(path, "rb"),
            as_attachment=True,
            filename="students_export.xlsx",
        )


# =========================================================
# FACULTY TEMPLATE DOWNLOAD
# =========================================================
class FacultyExcelTemplateDownload(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        context = _selected_faculty_context(request)
        missing = _validate_faculty_context_fields(context)
        if missing:
            return Response(
                {"error": "Required query params missing: " + ", ".join(missing)},
                status=400,
            )

        path = os.path.join(tempfile.gettempdir(), "faculty_template.xlsx")
        generate_faculty_template(
            path,
            metadata={
                "College Name": context["college_name"] or "SVR Engineering College",
                "Department": context["department"],
                "Designation": context["designation"],
                "Template Version": FACULTY_TEMPLATE_VERSION,
            },
        )
        return FileResponse(
            open(path, "rb"),
            as_attachment=True,
            filename="faculty_template.xlsx",
        )


# =========================================================
# FACULTY EXCEL UPLOAD (USERNAME + PASSWORD)
# =========================================================
class FacultyExcelUpload(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        context = _selected_faculty_context(request)
        missing = _validate_faculty_context_fields(context)
        if missing:
            return Response(
                {"error": "Please select and send: " + ", ".join(missing)},
                status=400,
            )

        file = request.FILES.get("file")
        if not file:
            return Response({"error": "Excel file required"}, status=400)

        try:
            metadata = _read_faculty_sheet_metadata(file)
        except Exception:
            return Response(
                {"error": "Invalid faculty template. Could not read metadata rows."},
                status=400,
            )

        if metadata.get("template_version") != FACULTY_TEMPLATE_VERSION:
            return Response(
                {
                    "error": (
                        "Template version mismatch. "
                        "Please download a fresh faculty template and upload again."
                    )
                },
                status=400,
            )

        mismatches = []
        if metadata.get("college_name", "").lower() != context["college_name"].lower():
            mismatches.append("college_name")
        if metadata.get("department", "").lower() != context["department"].lower():
            mismatches.append("department")
        if metadata.get("designation", "").lower() != context["designation"].lower():
            mismatches.append("designation")

        if mismatches:
            return Response(
                {
                    "error": (
                        "Selected filters do not match template metadata: "
                        + ", ".join(mismatches)
                    ),
                    "selected": context,
                    "template_metadata": metadata,
                },
                status=400,
            )

        file.seek(0)
        df = pd.read_excel(file, header=len(FACULTY_TEMPLATE_METADATA_ROWS), dtype=str)

        if list(df.columns) != FACULTY_EXCEL_COLUMNS:
            return Response(
                {"error": "Do not change column names or order"},
                status=400,
            )

        created_count = 0
        updated_count = 0
        skipped_count = 0
        row_errors = []
        row_warnings = []

        with transaction.atomic():
            for row_idx, r in df.iterrows():
                excel_row = int(row_idx) + 2
                id_no = _normalize_faculty_id(r.get("Employee ID No"))
                if not id_no:
                    skipped_count += 1
                    continue

                email = f"{id_no.lower()}@svrec.ac.in"

                # Resolve user safely by username/email to avoid unique-key crashes.
                user_by_username = User.objects.filter(username=id_no).first()
                user_by_email = User.objects.filter(email=email).first()

                # If both exist but are different users, prefer username match to keep login stable.
                # Keep that user's current email to avoid unique-email collisions.
                if user_by_username and user_by_email and user_by_username.id != user_by_email.id:
                    user = user_by_username
                    row_warnings.append({
                        "row": excel_row,
                        "id_no": id_no,
                        "message": "Username matched existing user with different email owner; kept existing username record",
                    })
                else:
                    user = user_by_username or user_by_email
                user_was_created = False

                try:
                    if not user:
                        user = User.objects.create_user(
                            email=email,
                            password=id_no,
                            username=id_no,
                            role="FACULTY",
                            is_active=True,
                            is_staff=True,
                        )
                        user_was_created = True
                    else:
                        user_updates = []
                        if user.username != id_no:
                            user.username = id_no
                            user_updates.append("username")
                        # Update email only if no different account already owns it.
                        if user.email != email:
                            email_owner = User.objects.filter(email=email).exclude(id=user.id).first()
                            if email_owner:
                                row_warnings.append({
                                    "row": excel_row,
                                    "id_no": id_no,
                                    "message": "Target email already used by another account; kept existing user email",
                                })
                            else:
                                user.email = email
                                user_updates.append("email")
                        if user.role != "FACULTY":
                            user.role = "FACULTY"
                            user_updates.append("role")
                        if not user.is_staff:
                            user.is_staff = True
                            user_updates.append("is_staff")
                        if not user.is_active:
                            user.is_active = True
                            user_updates.append("is_active")
                        if user_updates:
                            user.save(update_fields=user_updates)
                except IntegrityError as exc:
                    row_errors.append({
                        "row": excel_row,
                        "id_no": id_no,
                        "message": f"User conflict: {str(exc)}",
                    })
                    continue

                try:
                    faculty, faculty_created = Faculty.objects.update_or_create(
                        id_no=id_no,
                        defaults={
                            "user": user,
                            "faculty_id": id_no,
                            "full_name": _clean_text(r.get("Full Name as per SSC", "")),
                            "gender": _clean_text(r.get("Gender", "")),
                            "dob": parse_date(r.get("DOB")),
                            "joining_date": parse_date(r.get("Joining Date")),
                            "relieving_date": parse_date(r.get("Relieving Date")),
                            "salary": parse_decimal(r.get("Salary")),
                            "qualification": _clean_text(r.get("Highest Qualification", "")),
                            "reference_name": _clean_text(r.get("Reference Name", "")),
                            "husband_wife_name": _clean_text(r.get("Husband/Wife Name", "")),
                            "mother_name": _clean_text(r.get("Mother Name", "")),
                            "father_name": _clean_text(r.get("Father Name", "")),
                            "nationality": _clean_text(r.get("Nationality", "")),
                            "religion": _clean_text(r.get("Religion", "")),
                            "wedding_date": parse_date(r.get("Wedding Date")),
                            "caste": _clean_text(r.get("Caste", "")),
                            "reservation_category": _clean_text(r.get("Reservation Category", "")),
                            "minority_indicator": _clean_text(r.get("Minority Indicator", "")),
                            "mobile_no": _clean_text(r.get("Mobile No", "")),
                            "other_mobile_no": _clean_text(r.get("Other Mobile No", "")),
                            "email": email,
                            "is_physically_challenged": _clean_text(r.get("Is Physically Challenged", "")),
                            "blood_group": _clean_text(r.get("Blood Group", "")),
                            "aadhar_no": _clean_text(r.get("Aadhar No", "")),
                            "pan_number": _clean_text(r.get("PAN Number", "")),
                            "aicte_id": _clean_text(r.get("AICTE ID", "")),
                            "licence_number": _clean_text(r.get("Licence Number", "")),
                            "department": _clean_text(r.get("Department", "")),
                            "designation": _clean_text(r.get("Designation", "")),
                            "d_no": _clean_text(r.get("D.No", "")),
                            "street": _clean_text(r.get("Street", "")),
                            "village": _clean_text(r.get("Village", "")),
                            "district": _clean_text(r.get("District", "")),
                            "pincode": _clean_text(r.get("Pincode", "")),
                            "area": _clean_text(r.get("Area", "")),
                            "address": _build_address(
                                r.get("D.No", ""),
                                r.get("Street", ""),
                                r.get("Village", ""),
                                r.get("District", ""),
                                r.get("Pincode", ""),
                                r.get("Area", ""),
                            ),
                        },
                    )
                    if user_was_created or faculty_created:
                        created_count += 1
                    else:
                        updated_count += 1
                except IntegrityError as exc:
                    row_errors.append({
                        "row": excel_row,
                        "id_no": id_no,
                        "message": f"Faculty conflict: {str(exc)}",
                    })
                    continue

        return Response(
            {
                "success": len(row_errors) == 0,
                "message": (
                    "Faculty Excel processed successfully"
                    if len(row_errors) == 0
                    else "Faculty Excel processed with some row issues"
                ),
                "summary": {
                    "created": created_count,
                    "updated": updated_count,
                    "skipped": skipped_count,
                    "errors": len(row_errors),
                    "warnings": len(row_warnings),
                },
                "errors": row_errors[:20],
                "warnings": row_warnings[:20],
            },
            status=status.HTTP_200_OK,
        )


# =========================================================
# FACULTY EXCEL EXPORT
# =========================================================
class FacultyExcelExport(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        data = []
        for i, f in enumerate(Faculty.objects.all(), start=1):
            d_no = (f.d_no or "").strip()
            street = (f.street or "").strip()
            village = (f.village or "").strip()
            district = (f.district or "").strip()
            pincode = (f.pincode or "").strip()
            area = (f.area or "").strip()

            if not any([d_no, street, village, district, pincode, area]) and (f.address or "").strip():
                parts = [p.strip() for p in f.address.split("|")]
                if len(parts) >= 6:
                    d_no, street, village, district, pincode, area = parts[:6]

            data.append({
                "Sno": i,
                "Employee ID No": f.id_no,
                "Full Name as per SSC": f.full_name,
                "Gender": f.gender,
                "DOB": f.dob.strftime("%d-%m-%Y") if f.dob else "",
                "Joining Date": f.joining_date.strftime("%d-%m-%Y") if f.joining_date else "",
                "Relieving Date": f.relieving_date.strftime("%d-%m-%Y") if f.relieving_date else "",
                "Salary": str(f.salary) if f.salary is not None else "",
                "Highest Qualification": f.qualification,
                "Reference Name": f.reference_name,
                "Husband/Wife Name": f.husband_wife_name,
                "Mother Name": f.mother_name,
                "Father Name": f.father_name,
                "Nationality": f.nationality,
                "Religion": f.religion,
                "Wedding Date": f.wedding_date.strftime("%d-%m-%Y") if f.wedding_date else "",
                "Caste": f.caste,
                "Reservation Category": f.reservation_category,
                "Minority Indicator": f.minority_indicator,
                "Mobile No": f.mobile_no,
                "Other Mobile No": f.other_mobile_no,
                "Email": f.email,
                "Is Physically Challenged": f.is_physically_challenged,
                "Blood Group": f.blood_group,
                "Aadhar No": f.aadhar_no,
                "PAN Number": f.pan_number,
                "AICTE ID": f.aicte_id,
                "Licence Number": f.licence_number,
                "Department": f.department,
                "Designation": f.designation,
                "D.No": d_no,
                "Street": street,
                "Village": village,
                "District": district,
                "Pincode": pincode,
                "Area": area,
            })

        df = pd.DataFrame(data, columns=FACULTY_EXCEL_COLUMNS)
        path = os.path.join(tempfile.gettempdir(), "faculty_export.xlsx")
        df.to_excel(path, index=False)

        return FileResponse(
            open(path, "rb"),
            as_attachment=True,
            filename="faculty_export.xlsx",
        )


class AdminCreateFacultyView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        admin_user = _resolve_admin_user(request)

        # If a user session exists and it's not admin, block.
        # If no authenticated session exists, allow creation for current frontend flow.
        if getattr(request, "user", None) and request.user.is_authenticated:
            if getattr(request.user, "role", None) != "ADMIN":
                return Response(
                    {"success": False, "message": "Only admin users can create faculty accounts"},
                    status=status.HTTP_403_FORBIDDEN,
                )

        id_no = _normalize_faculty_id(request.data.get("id_no") or request.data.get("username"))
        password = request.data.get("password")

        if not id_no:
            return Response(
                {"success": False, "message": "Employee ID No is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not password:
            return Response(
                {"success": False, "message": "Password is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if len(str(password)) < 6:
            return Response(
                {"success": False, "message": "Password must be at least 6 characters"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        email = (request.data.get("email") or f"{id_no.lower()}@svrec.ac.in").strip().lower()
        try:
            validate_email(email)
        except ValidationError:
            return Response(
                {"success": False, "message": "Invalid email format"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if User.objects.filter(username=id_no).exists():
            return Response(
                {"success": False, "message": "Faculty username already exists"},
                status=status.HTTP_409_CONFLICT,
            )

        if Faculty.objects.filter(id_no=id_no).exists():
            return Response(
                {"success": False, "message": "Faculty Employee ID No already exists"},
                status=status.HTTP_409_CONFLICT,
            )

        if User.objects.filter(email=email).exists():
            return Response(
                {"success": False, "message": "Email already mapped to another account"},
                status=status.HTTP_409_CONFLICT,
            )

        try:
            with transaction.atomic():
                faculty_user = User.objects.create_user(
                    email=email,
                    password=password,
                    username=id_no,
                    role="FACULTY",
                    is_active=True,
                    is_staff=True,
                )

                faculty = Faculty.objects.create(
                    user=faculty_user,
                    id_no=id_no,
                    faculty_id=id_no,
                    full_name=(request.data.get("full_name") or "").strip(),
                    email=email,
                    mobile_no=(request.data.get("mobile_no") or "").strip(),
                    department=(request.data.get("department") or "").strip(),
                    designation=(request.data.get("designation") or "").strip(),
                    address=(request.data.get("address") or "").strip(),
                )
        except IntegrityError:
            return Response(
                {"success": False, "message": "Unable to create faculty due to duplicate data"},
                status=status.HTTP_409_CONFLICT,
            )

        return Response(
            {
                "success": True,
                "message": "Faculty account created successfully",
                "data": {
                    "user_id": faculty_user.id,
                    "role": faculty_user.role,
                    "username": faculty_user.username,
                    "id_no": faculty.id_no,
                    "email": faculty_user.email,
                    "created_by_admin": (
                        (admin_user.username or admin_user.email)
                        if admin_user
                        else "unverified-session"
                    ),
                },
            },
            status=status.HTTP_201_CREATED,
        )


class FacultyListView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        q = (request.query_params.get("q") or "").strip()
        department = (request.query_params.get("department") or "").strip()
        designation = (request.query_params.get("designation") or "").strip()
        status_filter = (request.query_params.get("status") or "active").strip().lower()

        qs = Faculty.objects.all().order_by("id_no")
        if q:
            qs = qs.filter(
                Q(id_no__icontains=q)
                | Q(full_name__icontains=q)
                | Q(email__icontains=q)
                | Q(mobile_no__icontains=q)
                | Q(department__icontains=q)
                | Q(designation__icontains=q)
            )
        today = timezone.now().date()
        if status_filter == "relieved":
            qs = qs.filter(relieving_date__isnull=False, relieving_date__lte=today)
        elif status_filter == "all":
            pass
        else:
            qs = qs.filter(Q(relieving_date__isnull=True) | Q(relieving_date__gt=today))

        rows_source = []
        for f in qs:
            if department and not _matches_compact(f.department, department):
                continue
            if designation and not _matches_compact(f.designation, designation):
                continue
            rows_source.append(f)

        rows = []
        for f in rows_source[:500]:
            rows.append(
                {
                    "id_no": f.id_no,
                    "full_name": f.full_name,
                    "gender": f.gender,
                    "email": f.email,
                    "mobile_no": f.mobile_no,
                    "department": f.department,
                    "designation": f.designation,
                    "joining_date": f.joining_date.strftime("%Y-%m-%d") if f.joining_date else "",
                    "relieving_date": f.relieving_date.strftime("%Y-%m-%d") if f.relieving_date else "",
                    "qualification": f.qualification,
                    "reference_name": f.reference_name,
                }
            )

        return Response({"success": True, "count": len(rows), "results": rows}, status=200)


class FacultyRelievedListView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        q = (request.query_params.get("q") or "").strip()
        department = (request.query_params.get("department") or "").strip()
        designation = (request.query_params.get("designation") or "").strip()
        today = timezone.now().date()

        qs = Faculty.objects.filter(relieving_date__isnull=False, relieving_date__lte=today).order_by("-relieving_date", "id_no")
        if q:
            qs = qs.filter(
                Q(id_no__icontains=q)
                | Q(full_name__icontains=q)
                | Q(email__icontains=q)
                | Q(mobile_no__icontains=q)
                | Q(department__icontains=q)
                | Q(designation__icontains=q)
            )
        rows_source = []
        for f in qs:
            if department and not _matches_compact(f.department, department):
                continue
            if designation and not _matches_compact(f.designation, designation):
                continue
            rows_source.append(f)

        rows = []
        for f in rows_source[:500]:
            rows.append(
                {
                    "id_no": f.id_no,
                    "full_name": f.full_name,
                    "department": f.department,
                    "designation": f.designation,
                    "relieving_date": f.relieving_date.strftime("%Y-%m-%d") if f.relieving_date else "",
                    "mobile_no": f.mobile_no,
                    "email": f.email,
                }
            )

        return Response({"success": True, "count": len(rows), "results": rows}, status=200)


class FacultySingleSaveView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        id_no = _normalize_faculty_id(request.data.get("id_no") or request.data.get("username"))
        if not id_no:
            return Response({"success": False, "message": "Employee ID No is required"}, status=400)

        full_name = (request.data.get("full_name") or "").strip()
        department = (request.data.get("department") or "").strip()
        designation = (request.data.get("designation") or "").strip()
        email = (request.data.get("email") or f"{id_no.lower()}@svrec.ac.in").strip().lower()
        mobile_no = (request.data.get("mobile_no") or "").strip()
        password = request.data.get("password")

        if not full_name:
            return Response({"success": False, "message": "Full Name as per SSC is required"}, status=400)
        if not department:
            return Response({"success": False, "message": "Department is required"}, status=400)
        if not designation:
            return Response({"success": False, "message": "Designation is required"}, status=400)
        try:
            validate_email(email)
        except ValidationError:
            return Response({"success": False, "message": "Invalid email format"}, status=400)

        existing_faculty = Faculty.objects.filter(id_no=id_no).first()
        if existing_faculty:
            user = existing_faculty.user
            if user and user.email != email:
                email_owner = User.objects.filter(email=email).exclude(id=user.id).first()
                if email_owner:
                    return Response({"success": False, "message": "Email already mapped to another account"}, status=409)
            with transaction.atomic():
                if user:
                    changed = []
                    if user.email != email:
                        user.email = email
                        changed.append("email")
                    if user.username != id_no:
                        user.username = id_no
                        changed.append("username")
                    if changed:
                        user.save(update_fields=changed)

                existing_faculty.full_name = full_name
                existing_faculty.gender = (request.data.get("gender") or "").strip()
                existing_faculty.dob = parse_date(request.data.get("dob"))
                existing_faculty.joining_date = parse_date(request.data.get("joining_date"))
                existing_faculty.relieving_date = parse_date(request.data.get("relieving_date"))
                existing_faculty.salary = parse_decimal(request.data.get("salary"))
                existing_faculty.qualification = (request.data.get("qualification") or "").strip()
                existing_faculty.blood_group = (request.data.get("blood_group") or "").strip()
                existing_faculty.reference_name = (request.data.get("reference_name") or "").strip()
                existing_faculty.husband_wife_name = (request.data.get("husband_wife_name") or "").strip()
                existing_faculty.aadhar_no = (request.data.get("aadhar_no") or "").strip()
                existing_faculty.pan_number = (request.data.get("pan_number") or "").strip()
                existing_faculty.aicte_id = (request.data.get("aicte_id") or "").strip()
                existing_faculty.mobile_no = mobile_no
                existing_faculty.other_mobile_no = (request.data.get("other_mobile_no") or "").strip()
                existing_faculty.email = email
                existing_faculty.department = department
                existing_faculty.designation = designation
                existing_faculty.d_no = (request.data.get("d_no") or "").strip()
                existing_faculty.street = (request.data.get("street") or "").strip()
                existing_faculty.village = (request.data.get("village") or "").strip()
                existing_faculty.district = (request.data.get("district") or "").strip()
                existing_faculty.pincode = (request.data.get("pincode") or "").strip()
                existing_faculty.area = (request.data.get("area") or "").strip()
                existing_faculty.address = " | ".join(
                    [
                        x
                        for x in [
                            existing_faculty.d_no,
                            existing_faculty.street,
                            existing_faculty.village,
                            existing_faculty.district,
                            existing_faculty.pincode,
                            existing_faculty.area,
                        ]
                        if x
                    ]
                )
                existing_faculty.save()

            return Response({"success": True, "message": "Faculty updated successfully", "id_no": id_no}, status=200)

        if not password:
            return Response({"success": False, "message": "Password is required for new faculty"}, status=400)
        if len(str(password)) < 6:
            return Response({"success": False, "message": "Password must be at least 6 characters"}, status=400)
        if User.objects.filter(username=id_no).exists():
            return Response({"success": False, "message": "Faculty username already exists"}, status=409)
        if User.objects.filter(email=email).exists():
            return Response({"success": False, "message": "Email already mapped to another account"}, status=409)

        try:
            with transaction.atomic():
                faculty_user = User.objects.create_user(
                    email=email,
                    password=password,
                    username=id_no,
                    role="FACULTY",
                    is_active=True,
                    is_staff=True,
                )
                Faculty.objects.create(
                    user=faculty_user,
                    id_no=id_no,
                    faculty_id=id_no,
                    full_name=full_name,
                    gender=(request.data.get("gender") or "").strip(),
                    dob=parse_date(request.data.get("dob")),
                    joining_date=parse_date(request.data.get("joining_date")),
                    relieving_date=parse_date(request.data.get("relieving_date")),
                    salary=parse_decimal(request.data.get("salary")),
                    qualification=(request.data.get("qualification") or "").strip(),
                    blood_group=(request.data.get("blood_group") or "").strip(),
                    reference_name=(request.data.get("reference_name") or "").strip(),
                    husband_wife_name=(request.data.get("husband_wife_name") or "").strip(),
                    aadhar_no=(request.data.get("aadhar_no") or "").strip(),
                    pan_number=(request.data.get("pan_number") or "").strip(),
                    aicte_id=(request.data.get("aicte_id") or "").strip(),
                    email=email,
                    mobile_no=mobile_no,
                    other_mobile_no=(request.data.get("other_mobile_no") or "").strip(),
                    department=department,
                    designation=designation,
                    d_no=(request.data.get("d_no") or "").strip(),
                    street=(request.data.get("street") or "").strip(),
                    village=(request.data.get("village") or "").strip(),
                    district=(request.data.get("district") or "").strip(),
                    pincode=(request.data.get("pincode") or "").strip(),
                    area=(request.data.get("area") or "").strip(),
                    address=" | ".join(
                        [
                            str(x).strip()
                            for x in [
                                request.data.get("d_no"),
                                request.data.get("street"),
                                request.data.get("village"),
                                request.data.get("district"),
                                request.data.get("pincode"),
                                request.data.get("area"),
                            ]
                            if str(x or "").strip()
                        ]
                    ),
                )
        except IntegrityError:
            return Response({"success": False, "message": "Unable to create faculty due to duplicate data"}, status=409)

        return Response({"success": True, "message": "Faculty created successfully", "id_no": id_no}, status=201)


class FacultyTransferDepartmentView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        from_department = (request.data.get("from_department") or "").strip()
        from_designation = (request.data.get("from_designation") or "").strip()
        to_department = (request.data.get("to_department") or "").strip()
        to_designation = (request.data.get("to_designation") or "").strip()
        include_relieved = bool(request.data.get("include_relieved", False))

        if not from_department:
            return Response({"success": False, "message": "From department is required"}, status=400)
        if not to_department:
            return Response({"success": False, "message": "To department is required"}, status=400)

        qs = Faculty.objects.filter(department__iexact=from_department)
        if from_designation:
            qs = qs.filter(designation__iexact=from_designation)
        if not include_relieved:
            today = timezone.now().date()
            qs = qs.filter(Q(relieving_date__isnull=True) | Q(relieving_date__gt=today))

        matched = qs.count()
        if matched == 0:
            return Response(
                {"success": True, "message": "No faculty found for selected source filters", "summary": {"matched": 0, "updated": 0}},
                status=200,
            )

        updates = {"department": to_department}
        if to_designation:
            updates["designation"] = to_designation
        updated = qs.update(**updates)

        return Response(
            {
                "success": True,
                "message": "Faculty transfer completed successfully",
                "summary": {"matched": matched, "updated": updated},
                "from": {"department": from_department, "designation": from_designation},
                "to": {"department": to_department, "designation": to_designation},
            },
            status=200,
        )


class FacultyDetailView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, id_no):
        fid = _normalize_faculty_id(id_no)
        f = Faculty.objects.filter(id_no=fid).first()
        if not f:
            return Response({"success": False, "message": "Faculty not found"}, status=404)

        data = {
            "Employee ID No": f.id_no,
            "Full Name as per SSC": f.full_name,
            "Gender": f.gender,
            "DOB": f.dob.strftime("%Y-%m-%d") if f.dob else "",
            "Joining Date": f.joining_date.strftime("%Y-%m-%d") if f.joining_date else "",
            "Relieving Date": f.relieving_date.strftime("%Y-%m-%d") if f.relieving_date else "",
            "Salary": str(f.salary) if f.salary is not None else "",
            "Highest Qualification": f.qualification,
            "Blood Group": f.blood_group,
            "Aadhar No": f.aadhar_no,
            "PAN Number": f.pan_number,
            "AICTE ID": f.aicte_id,
            "Reference Name": f.reference_name,
            "Husband/Wife Name": f.husband_wife_name,
            "Mobile No": f.mobile_no,
            "Other Mobile No": f.other_mobile_no,
            "Email": f.email,
            "Department": f.department,
            "Designation": f.designation,
            "D.No": f.d_no,
            "Street": f.street,
            "Village": f.village,
            "District": f.district,
            "Pincode": f.pincode,
            "Area": f.area,
        }
        return Response({"success": True, "data": data}, status=200)

    def put(self, request, id_no):
        fid = _normalize_faculty_id(id_no)
        f = Faculty.objects.filter(id_no=fid).first()
        if not f:
            return Response({"success": False, "message": "Faculty not found"}, status=404)

        payload = request.data
        full_name = (payload.get("full_name") or payload.get("Full Name as per SSC") or "").strip()
        if not full_name:
            return Response({"success": False, "message": "Full Name as per SSC is required"}, status=400)

        department = (payload.get("department") or payload.get("Department") or "").strip()
        designation = (payload.get("designation") or payload.get("Designation") or "").strip()
        if not department:
            return Response({"success": False, "message": "Department is required"}, status=400)
        if not designation:
            return Response({"success": False, "message": "Designation is required"}, status=400)

        email = (payload.get("email") or payload.get("Email") or f.email or f"{fid.lower()}@svrec.ac.in").strip().lower()
        try:
            validate_email(email)
        except ValidationError:
            return Response({"success": False, "message": "Invalid email format"}, status=400)

        user = f.user
        with transaction.atomic():
            if user:
                changed = []
                if user.email != email:
                    email_owner = User.objects.filter(email=email).exclude(id=user.id).first()
                    if email_owner:
                        return Response({"success": False, "message": "Email already mapped to another account"}, status=409)
                    user.email = email
                    changed.append("email")
                if user.username != fid:
                    user.username = fid
                    changed.append("username")
                if changed:
                    user.save(update_fields=changed)

            f.full_name = full_name
            f.gender = (payload.get("gender") or payload.get("Gender") or "").strip()
            f.dob = parse_date(payload.get("dob") or payload.get("DOB"))
            f.joining_date = parse_date(payload.get("joining_date") or payload.get("Joining Date"))
            f.relieving_date = parse_date(payload.get("relieving_date") or payload.get("Relieving Date"))
            f.salary = parse_decimal(payload.get("salary") or payload.get("Salary"))
            f.qualification = (payload.get("qualification") or payload.get("Highest Qualification") or "").strip()
            f.blood_group = (payload.get("blood_group") or payload.get("Blood Group") or "").strip()
            f.aadhar_no = (payload.get("aadhar_no") or payload.get("Aadhar No") or "").strip()
            f.pan_number = (payload.get("pan_number") or payload.get("PAN Number") or "").strip()
            f.aicte_id = (payload.get("aicte_id") or payload.get("AICTE ID") or "").strip()
            f.reference_name = (payload.get("reference_name") or payload.get("Reference Name") or "").strip()
            f.husband_wife_name = (payload.get("husband_wife_name") or payload.get("Husband/Wife Name") or "").strip()
            f.mobile_no = (payload.get("mobile_no") or payload.get("Mobile No") or "").strip()
            f.other_mobile_no = (payload.get("other_mobile_no") or payload.get("Other Mobile No") or "").strip()
            f.email = email
            f.department = department
            f.designation = designation
            f.d_no = (payload.get("d_no") or payload.get("D.No") or "").strip()
            f.street = (payload.get("street") or payload.get("Street") or "").strip()
            f.village = (payload.get("village") or payload.get("Village") or "").strip()
            f.district = (payload.get("district") or payload.get("District") or "").strip()
            f.pincode = (payload.get("pincode") or payload.get("Pincode") or "").strip()
            f.area = (payload.get("area") or payload.get("Area") or "").strip()
            f.address = " | ".join([x for x in [f.d_no, f.street, f.village, f.district, f.pincode, f.area] if x])
            f.save()

        return Response({"success": True, "message": "Faculty updated successfully", "id_no": fid}, status=200)


class FacultyRelieveView(APIView):
    permission_classes = [AllowAny]

    def post(self, request, id_no):
        fid = _normalize_faculty_id(id_no)
        f = Faculty.objects.filter(id_no=fid).first()
        if not f:
            return Response({"success": False, "message": "Faculty not found"}, status=404)

        relieving_date = parse_date(request.data.get("relieving_date")) or timezone.now().date()
        f.relieving_date = relieving_date
        f.save(update_fields=["relieving_date", "updated_at"])
        return Response(
            {
                "success": True,
                "message": "Faculty relieved successfully",
                "id_no": fid,
                "relieving_date": relieving_date.strftime("%Y-%m-%d"),
            },
            status=200,
        )


class FacultyRemoveView(APIView):
    permission_classes = [AllowAny]

    def delete(self, request, id_no):
        fid = _normalize_faculty_id(id_no)
        f = Faculty.objects.filter(id_no=fid).first()
        if not f:
            return Response({"success": False, "message": "Faculty not found"}, status=404)

        user = f.user
        with transaction.atomic():
            f.delete()
            if user:
                user.is_active = False
                user.save(update_fields=["is_active"])

        return Response({"success": True, "message": "Faculty removed successfully", "id_no": fid}, status=200)


class AdminUserPasswordUpdateView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        identifier = (request.data.get("username") or request.data.get("identifier") or "").strip()
        new_password = (request.data.get("new_password") or request.data.get("password") or "").strip()

        if not identifier:
            return Response({"success": False, "message": "Username is required"}, status=400)
        if not new_password:
            return Response({"success": False, "message": "New password is required"}, status=400)
        if len(new_password) < 6:
            return Response({"success": False, "message": "Password must be at least 6 characters"}, status=400)

        user = User.objects.filter(username=identifier).first()
        if not user:
            return Response({"success": False, "message": "Username not found"}, status=404)

        user.set_password(new_password)
        user.save(update_fields=["password"])
        return Response(
            {
                "success": True,
                "message": "Password updated successfully",
                "data": {"username": user.username, "role": user.role},
            },
            status=200,
        )
