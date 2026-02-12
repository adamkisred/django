from collections import defaultdict
from datetime import datetime, time as dt_time
from decimal import Decimal, InvalidOperation
import os
import tempfile

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.protection import Protection
from openpyxl.worksheet.datavalidation import DataValidation

from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView

from django.db import transaction
from django.db.models import Q
from django.http import FileResponse
from django.utils import timezone

from academics.models import Subject, TimetableMapping, TimeSlot, Timetable, MidMark, SubjectFacultyMapping
from academics.services.timetable_auto import auto_generate_timetable, TimetableGenerationError
from faculty.models import Faculty
from students.models import Student


def _clean(value):
    return str(value or "").strip()


DEFAULT_SECTIONS = ["A", "B", "C", "D"]
PERIOD_MASTER = {
    1: ("1st Period", "9:30 AM - 10:20 AM"),
    2: ("2nd Period", "10:20 AM - 11:10 AM"),
    3: ("3rd Period", "11:30 AM - 12:20 PM"),
    4: ("4th Period", "12:20 PM - 1:30 PM"),
    5: ("5th Period", "2:00 PM - 2:50 PM"),
    6: ("6th Period", "2:50 PM - 3:40 PM"),
    7: ("7th Period", "3:40 PM - 4:30 PM"),
}
VALID_DAYS = {"Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"}
VALID_SUBJECT_TYPES = {"THEORY", "PRACTICAL", "CRT", "MENTORING", "OTHER"}
DEFAULT_REGULATIONS = ["R18", "R20", "R23", "R24"]
DEFAULT_EXAM_TYPES = ["MID-1", "MID-2", "MID-3"]
MAX_DAILY_FACULTY_WORKLOAD = 6
MAX_WEEKLY_FACULTY_WORKLOAD = 30
BREAK_TRANSITIONS = {(2, 3), (4, 5)}
VALID_PRACTICAL_BLOCKS = {(1, 2, 3), (2, 3, 4), (5, 6, 7)}
DAY_SHORT_TO_FULL = {
    "MON": "Monday",
    "TUE": "Tuesday",
    "WED": "Wednesday",
    "THU": "Thursday",
    "FRI": "Friday",
    "SAT": "Saturday",
}
DAY_FULL_TO_SHORT = {v: k for k, v in DAY_SHORT_TO_FULL.items()}
SUBJECT_TEMPLATE_METADATA_ROWS = [
    "College Name",
    "Branch",
    "Semester",
    "Curriculum",
    "Curid",
    "Academic Year",
]
SUBJECT_EXCEL_COLUMNS = [
    "SubOrder",
    "SubCode",
    "RefCode",
    "IntExamCode",
    "ExtExamCode",
    "Name",
    "RunningCur",
    "Elective(0-No,1-YES)",
    "Elective SubName",
    "Replacement(0-No,1-Yes)",
    "IntMax",
    "ExtMax",
    "Subtype(0-Theory,1-Practical,2-Drawing,3-Project,4-Others)",
    "Credits",
    "SubjectCategory(Theory/Practical/Mentoring/CRT/Other)",
]

MID_TEMPLATE_METADATA_ROWS = [
    "College Name",
    "Branch",
    "Section",
    "Semester",
    "Academic Year",
]


def _subject_context_from_request(request):
    return {
        "college_name": _clean(request.data.get("college_name") or request.query_params.get("college_name") or "SVR Engineering College"),
        "academic_year": _clean(request.data.get("academic_year") or request.query_params.get("academic_year")),
        "branch": _clean(request.data.get("branch") or request.query_params.get("branch")),
        "semester": _clean(request.data.get("semester") or request.query_params.get("semester")),
        "regulation": _clean(request.data.get("regulation") or request.query_params.get("regulation") or "R20").upper(),
        "curid": _clean(request.data.get("curid") or request.query_params.get("curid") or "3"),
    }


def _subject_context_missing(context):
    missing = []
    if not context["academic_year"]:
        missing.append("academic_year")
    if not context["branch"]:
        missing.append("branch")
    if not context["semester"]:
        missing.append("semester")
    if not context["regulation"]:
        missing.append("regulation")
    return missing


def _safe_text(value):
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def _to_full_day(value):
    day = _clean(value)
    if not day:
        return ""
    if day in VALID_DAYS:
        return day
    return DAY_SHORT_TO_FULL.get(day.upper(), day)


def _to_short_day(value):
    day = _clean(value)
    if not day:
        return ""
    if day.upper() in DAY_SHORT_TO_FULL:
        return day.upper()
    return DAY_FULL_TO_SHORT.get(day, day[:3].upper())


def _ensure_time_slots():
    slots = {}
    slot_times = {
        1: (dt_time(9, 30), dt_time(10, 20)),
        2: (dt_time(10, 20), dt_time(11, 10)),
        3: (dt_time(11, 30), dt_time(12, 20)),
        4: (dt_time(12, 20), dt_time(13, 10)),
        5: (dt_time(14, 0), dt_time(14, 50)),
        6: (dt_time(14, 50), dt_time(15, 40)),
        7: (dt_time(15, 40), dt_time(16, 30)),
    }
    for day in VALID_DAYS:
        for period, (start_t, end_t) in slot_times.items():
            obj, _ = TimeSlot.objects.get_or_create(
                day=day,
                period_number=period,
                defaults={"start_time": start_t, "end_time": end_t},
            )
            slots[(day, period)] = obj
    return slots


def _serialize_saved_timetable(qs):
    rows = []
    for t in qs.select_related("subject").order_by("day", "period_no"):
        rows.append(
            {
                "day": t.day,
                "day_short": _to_short_day(t.day),
                "period_no": t.period_no,
                "subject_id": t.subject.subject_id,
                "subject_code": t.subject.subject_id,
                "subject_name": t.subject.subject_name,
                "subject_type": t.subject.subject_type,
                "credits": str(t.subject.credits),
            }
        )
    return rows


def _resolve_faculty_ids_for_subject(context_key, subject_code, cache):
    cache_key = (*context_key, subject_code)
    if cache_key in cache:
        return cache[cache_key]
    faculty_ids = list(
        SubjectFacultyMapping.objects.filter(
            academic_year=context_key[0],
            branch=context_key[1],
            semester=context_key[2],
            section=context_key[3],
            subject__subject_id=subject_code,
        )
        .values_list("faculty_id", flat=True)
        .distinct()
    )
    cache[cache_key] = faculty_ids
    return faculty_ids


def _validate_faculty_conflicts_for_timetable_entries(
    *,
    academic_year,
    branch,
    semester,
    section,
    regulation,
    prepared,  # {(day, period): Subject}
):
    class_subject_map_qs = SubjectFacultyMapping.objects.filter(
        academic_year=academic_year,
        branch=branch,
        semester=semester,
        section=section,
    )
    subject_to_faculty = defaultdict(set)
    for row in class_subject_map_qs.select_related("subject"):
        subject_to_faculty[row.subject.subject_id].add(row.faculty_id)

    faculty_for_subject = {}
    for (day, period_no), subject in prepared.items():
        fids = subject_to_faculty.get(subject.subject_id, set())
        if not fids:
            return False, f"Faculty mapping not found for subject {subject.subject_id}. Save subject-faculty mapping first."
        if len(fids) > 1:
            return (
                False,
                f"Subject {subject.subject_id} is mapped to multiple faculties. Keep one faculty per subject.",
            )
        faculty_for_subject[(day, period_no)] = next(iter(fids))

    existing_busy = defaultdict(set)  # faculty_id -> {(day, period)}
    existing_day_load = defaultdict(int)  # (faculty_id, day) -> count
    existing_week_load = defaultdict(int)  # faculty_id -> count

    def _mark_existing(fid, day, period_no):
        slot = (day, period_no)
        if slot in existing_busy[fid]:
            return
        existing_busy[fid].add(slot)
        existing_day_load[(fid, day)] += 1
        existing_week_load[fid] += 1

    manual_qs = TimetableMapping.objects.exclude(
        academic_year=academic_year,
        branch=branch,
        semester=semester,
        section=section,
    )
    for row in manual_qs.values("faculty_id", "week_day", "period_no"):
        _mark_existing(row["faculty_id"], row["week_day"], row["period_no"])

    faculty_cache = {}
    auto_qs = Timetable.objects.exclude(
        academic_year=academic_year,
        branch=branch,
        semester=semester,
        section=section,
        regulation=regulation,
    ).select_related("subject")
    for row in auto_qs:
        context_key = (row.academic_year, row.branch, row.semester, row.section)
        subject_code = row.subject.subject_id
        for fid in _resolve_faculty_ids_for_subject(context_key, subject_code, faculty_cache):
            _mark_existing(fid, row.day, row.period_no)

    new_busy = defaultdict(set)
    new_day_load = defaultdict(int)
    new_week_load = defaultdict(int)

    for (day, period_no), fid in faculty_for_subject.items():
        if (day, period_no) in existing_busy[fid]:
            return (
                False,
                f"Faculty clash detected: faculty is already assigned on {day} period {period_no} in another class/section/semester.",
            )
        if (day, period_no) in new_busy[fid]:
            return (
                False,
                f"Faculty clash detected inside selected section: duplicate assignment on {day} period {period_no}.",
            )
        new_busy[fid].add((day, period_no))
        new_day_load[(fid, day)] += 1
        new_week_load[fid] += 1

        if existing_day_load[(fid, day)] + new_day_load[(fid, day)] > MAX_DAILY_FACULTY_WORKLOAD:
            return False, f"Daily faculty workload exceeded on {day}."
        if existing_week_load[fid] + new_week_load[fid] > MAX_WEEKLY_FACULTY_WORKLOAD:
            return False, "Weekly faculty workload exceeded."

    return True, ""


def _validate_practical_continuity_for_timetable_entries(prepared):
    # Practical subjects must be in continuous 3-period blocks.
    # They can cross tea break (2->3), but cannot cross lunch (4->5).
    # Valid blocks: (1,2,3), (2,3,4), (5,6,7).
    by_day_subject = defaultdict(list)  # (day, subject_code) -> [periods]
    for (day, period_no), subject in prepared.items():
        if subject.subject_type == "PRACTICAL":
            by_day_subject[(day, subject.subject_id)].append(period_no)

    for (day, subject_code), periods in by_day_subject.items():
        periods = sorted(periods)
        if len(periods) != 3:
            return (
                False,
                f"Practical subject {subject_code} on {day} must be exactly 3 continuous periods.",
            )
        if periods[1] != periods[0] + 1 or periods[2] != periods[1] + 1:
            return (
                False,
                f"Practical subject {subject_code} on {day} is not continuous.",
            )
        if tuple(periods) not in VALID_PRACTICAL_BLOCKS:
            return (
                False,
                f"Practical subject {subject_code} on {day} must be one of: (1,2,3), (2,3,4), (5,6,7).",
            )
    return True, ""


def _read_subject_sheet_metadata(file_obj):
    file_obj.seek(0)
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    metadata = {
        "college_name": _safe_text(ws.cell(row=1, column=2).value),
        "branch": _safe_text(ws.cell(row=2, column=2).value),
        "semester": _safe_text(ws.cell(row=3, column=2).value),
        "regulation": _safe_text(ws.cell(row=4, column=2).value).upper(),
        "curid": _safe_text(ws.cell(row=5, column=2).value),
        "academic_year": _safe_text(ws.cell(row=6, column=2).value),
    }
    wb.close()
    return metadata


def _build_subject_template(path, metadata):
    wb = Workbook()
    ws = wb.active
    ws.title = "Subjects"

    label_fill = PatternFill(fill_type="solid", start_color="EAF2FF", end_color="EAF2FF")
    header_fill = PatternFill(fill_type="solid", start_color="DDE7F3", end_color="DDE7F3")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for idx, key in enumerate(SUBJECT_TEMPLATE_METADATA_ROWS, start=1):
        ws.cell(row=idx, column=1, value=key).font = bold
        ws.cell(row=idx, column=1).fill = label_fill
        ws.cell(row=idx, column=2, value=metadata.get(key, ""))
        ws.cell(row=idx, column=1).protection = Protection(locked=True)
        ws.cell(row=idx, column=2).protection = Protection(locked=True)

    header_row = len(SUBJECT_TEMPLATE_METADATA_ROWS) + 1
    for col_idx, col_name in enumerate(SUBJECT_EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.protection = Protection(locked=True)

    for row_idx in range(header_row + 1, header_row + 201):
        for col_idx in range(1, len(SUBJECT_EXCEL_COLUMNS) + 1):
            ws.cell(row=row_idx, column=col_idx).protection = Protection(locked=False)

    # Subject category dropdown for easier and consistent type selection in template.
    category_col_idx = SUBJECT_EXCEL_COLUMNS.index("SubjectCategory(Theory/Practical/Mentoring/CRT/Other)") + 1
    category_col = chr(ord("A") + category_col_idx - 1)
    category_validation = DataValidation(
        type="list",
        formula1='"THEORY,PRACTICAL,MENTORING,CRT,OTHER"',
        allow_blank=True,
        showDropDown=True,
    )
    ws.add_data_validation(category_validation)
    category_validation.add(f"{category_col}{header_row + 1}:{category_col}{header_row + 200}")

    widths = [10, 16, 14, 16, 16, 36, 14, 16, 28, 18, 10, 10, 30, 10, 34]
    for idx, w in enumerate(widths, start=1):
        col = chr(ord("A") + idx - 1)
        ws.column_dimensions[col].width = w
    ws.freeze_panes = f"A{header_row + 1}"
    ws.protection.sheet = True
    ws.protection.enable()

    wb.save(path)
    wb.close()


class SubjectContextOptionsView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        now_year = datetime.now().year
        auto_years = [f"{y}-{y + 1}" for y in range(now_year - 2, now_year + 5)]
        db_years = list(
            Student.objects.exclude(academic_year="")
            .values_list("academic_year", flat=True)
            .distinct()
        )
        years = sorted(set(auto_years + db_years))
        branches = sorted(
            set(
                list(
                    Student.objects.exclude(branch="")
                    .values_list("branch", flat=True)
                    .distinct()
                )
                + ["CSE", "ECE", "EEE", "MECH", "CIVIL", "CSE (AI)", "IT"]
            )
        )
        semesters = ["1-1", "1-2", "2-1", "2-2", "3-1", "3-2", "4-1", "4-2"]
        regulations = sorted(
            set(
                DEFAULT_REGULATIONS
                + list(
                    Subject.objects.exclude(regulation="")
                    .values_list("regulation", flat=True)
                    .distinct()
                )
            )
        )
        return Response(
            {
                "success": True,
                "academic_years": years,
                "branches": branches,
                "semesters": semesters,
                "regulations": regulations,
            },
            status=200,
        )


class SubjectListCreateView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        context = _subject_context_from_request(request)
        missing = _subject_context_missing(context)
        if missing:
            return Response(
                {
                    "success": False,
                    "message": "Required params missing: " + ", ".join(missing),
                },
                status=400,
            )

        qs = Subject.objects.filter(
            academic_year=context["academic_year"],
            branch=context["branch"],
            semester=context["semester"],
            regulation=context["regulation"],
        ).order_by("subject_id")

        results = [
            {
                "id": s.id,
                "subject_code": s.subject_id,
                "subject_id": s.subject_id,
                "subject_name": s.subject_name,
                "subject_type": s.subject_type,
                "credits": str(s.credits),
                "regulation": s.regulation,
            }
            for s in qs
        ]
        return Response({"success": True, "count": len(results), "results": results}, status=200)

    def post(self, request):
        context = _subject_context_from_request(request)
        subject_id = _clean(request.data.get("subject_code") or request.data.get("subject_id")).upper()
        subject_name = _clean(request.data.get("subject_name"))
        subject_type = _clean(request.data.get("subject_type")).upper() or "THEORY"
        credits_raw = _clean(request.data.get("credits"))

        missing = _subject_context_missing(context)
        if missing:
            return Response(
                {
                    "success": False,
                    "message": "Required fields missing: " + ", ".join(missing),
                },
                status=400,
            )
        if not subject_id:
            return Response({"success": False, "message": "Subject Code is required"}, status=400)
        if not subject_name:
            return Response({"success": False, "message": "Subject Name is required"}, status=400)
        if subject_type not in VALID_SUBJECT_TYPES:
            return Response(
                {"success": False, "message": "subject_type must be THEORY, PRACTICAL, CRT, MENTORING or OTHER"},
                status=400,
            )
        try:
            credits = Decimal(credits_raw or "0")
        except (InvalidOperation, ValueError):
            return Response({"success": False, "message": "credits must be a valid number"}, status=400)
        if credits < 0:
            return Response({"success": False, "message": "credits must be 0 or higher"}, status=400)

        if Subject.objects.filter(
            academic_year=context["academic_year"],
            branch=context["branch"],
            semester=context["semester"],
            regulation=context["regulation"],
            subject_id=subject_id,
        ).exists():
            return Response(
                {"success": False, "message": "Subject Code already exists for selected context"},
                status=409,
            )

        if Subject.objects.filter(
            academic_year=context["academic_year"],
            branch=context["branch"],
            semester=context["semester"],
            regulation=context["regulation"],
            subject_name__iexact=subject_name,
        ).exists():
            return Response(
                {"success": False, "message": "Subject Name already exists for selected context"},
                status=409,
            )

        s = Subject.objects.create(
            academic_year=context["academic_year"],
            branch=context["branch"],
            semester=context["semester"],
            regulation=context["regulation"],
            subject_id=subject_id,
            subject_name=subject_name,
            subject_type=subject_type,
            credits=credits,
        )
        return Response(
            {
                "success": True,
                "message": "Subject added successfully",
                "data": {
                    "id": s.id,
                    "subject_code": s.subject_id,
                    "subject_id": s.subject_id,
                    "subject_name": s.subject_name,
                    "subject_type": s.subject_type,
                    "credits": str(s.credits),
                    "regulation": s.regulation,
                },
            },
            status=201,
        )


class SubjectDeleteView(APIView):
    permission_classes = [AllowAny]

    def delete(self, request, subject_pk):
        s = Subject.objects.filter(id=subject_pk).first()
        if not s:
            return Response({"success": False, "message": "Subject not found"}, status=404)
        s.delete()
        return Response({"success": True, "message": "Subject removed successfully"}, status=200)


class SubjectTemplateDownloadView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        context = _subject_context_from_request(request)
        missing = _subject_context_missing(context)
        if missing:
            return Response({"success": False, "message": "Required query params: " + ", ".join(missing)}, status=400)

        path = os.path.join(tempfile.gettempdir(), "subjects_template.xlsx")
        _build_subject_template(
            path,
            {
                "College Name": context["college_name"],
                "Branch": context["branch"],
                "Semester": context["semester"],
                "Curriculum": context["regulation"],
                "Curid": context["curid"],
                "Academic Year": context["academic_year"],
            },
        )
        return FileResponse(open(path, "rb"), as_attachment=True, filename="subjects_template.xlsx")


class SubjectExcelUploadView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        context = _subject_context_from_request(request)
        missing = _subject_context_missing(context)
        if missing:
            return Response({"success": False, "message": "Required fields: " + ", ".join(missing)}, status=400)

        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"success": False, "message": "Excel file is required"}, status=400)

        try:
            metadata = _read_subject_sheet_metadata(file_obj)
        except Exception:
            return Response({"success": False, "message": "Invalid template metadata. Download a fresh template."}, status=400)

        mismatches = []
        if metadata.get("college_name", "").lower() != context["college_name"].lower():
            mismatches.append("college_name")
        if metadata.get("branch", "").lower() != context["branch"].lower():
            mismatches.append("branch")
        if metadata.get("semester", "").lower() != context["semester"].lower():
            mismatches.append("semester")
        if metadata.get("regulation", "").lower() != context["regulation"].lower():
            mismatches.append("regulation")
        if metadata.get("curid", "").lower() != context["curid"].lower():
            mismatches.append("curid")
        if metadata.get("academic_year", "").lower() != context["academic_year"].lower():
            mismatches.append("academic_year")

        if mismatches:
            return Response(
                {
                    "success": False,
                    "message": "Selected filters do not match template metadata: " + ", ".join(mismatches),
                    "selected": context,
                    "template_metadata": metadata,
                },
                status=400,
            )

        file_obj.seek(0)
        df = pd.read_excel(file_obj, header=len(SUBJECT_TEMPLATE_METADATA_ROWS), dtype=str).fillna("")
        if list(df.columns) != SUBJECT_EXCEL_COLUMNS:
            return Response(
                {
                    "success": False,
                    "message": "Invalid header row. Do not edit protected column names.",
                    "expected": SUBJECT_EXCEL_COLUMNS,
                },
                status=400,
            )

        created = 0
        updated = 0
        skipped = 0
        errors = []

        with transaction.atomic():
            for idx, row in df.iterrows():
                excel_row = int(idx) + len(SUBJECT_TEMPLATE_METADATA_ROWS) + 2
                subject_code = _clean(row.get("SubCode")).upper()
                subject_name = _clean(row.get("Name"))
                subtype_raw = _clean(row.get("Subtype(0-Theory,1-Practical,2-Drawing,3-Project,4-Others)"))
                category_raw = _clean(row.get("SubjectCategory(Theory/Practical/Mentoring/CRT/Other)"))
                credits_raw = _clean(row.get("Credits"))

                if not subject_code and not subject_name:
                    skipped += 1
                    continue

                if not subject_code:
                    errors.append({"row": excel_row, "message": "Subject Code is required"})
                    continue
                if not subject_name:
                    errors.append({"row": excel_row, "message": "Subject Name is required"})
                    continue
                category_upper = category_raw.upper()
                if category_upper in {"THEORY", "PRACTICAL", "MENTORING", "CRT", "OTHER"}:
                    subject_type = category_upper
                else:
                    subtype_upper = subtype_raw.upper()
                    if not subtype_upper:
                        subject_type = "THEORY"
                    elif subtype_upper in {"0", "THEORY", "T"}:
                        subject_type = "THEORY"
                    elif subtype_upper in {"1", "2", "3", "PRACTICAL", "DRAWING", "PROJECT", "P"}:
                        subject_type = "PRACTICAL"
                    elif subtype_upper in {"CRT", "C"}:
                        subject_type = "CRT"
                    elif subtype_upper in {"MENTORING", "MENTOR"}:
                        subject_type = "MENTORING"
                    elif subtype_upper in {"4", "OTHER", "OTHERS"}:
                        subject_type = "OTHER"
                    else:
                        errors.append(
                            {
                                "row": excel_row,
                                "message": (
                                    "Invalid SubjectCategory. Use THEORY/PRACTICAL/MENTORING/CRT/OTHER "
                                    "or valid subtype fallback."
                                ),
                            }
                        )
                        continue
                try:
                    credits = Decimal(credits_raw or "0")
                except (InvalidOperation, ValueError):
                    errors.append({"row": excel_row, "message": "Credits must be numeric"})
                    continue

                obj, is_created = Subject.objects.update_or_create(
                    academic_year=context["academic_year"],
                    branch=context["branch"],
                    semester=context["semester"],
                    regulation=context["regulation"],
                    subject_id=subject_code,
                    defaults={
                        "subject_name": subject_name,
                        "subject_type": subject_type,
                        "credits": credits,
                    },
                )
                if is_created:
                    created += 1
                else:
                    updated += 1

        return Response(
            {
                "success": len(errors) == 0,
                "message": "Subject Excel processed successfully" if len(errors) == 0 else "Subject Excel processed with row errors",
                "summary": {"created": created, "updated": updated, "skipped": skipped, "errors": len(errors)},
                "errors": errors[:25],
            },
            status=200,
        )


class SubjectExcelExportView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        context = _subject_context_from_request(request)
        missing = _subject_context_missing(context)
        if missing:
            return Response({"success": False, "message": "Required query params: " + ", ".join(missing)}, status=400)

        path = os.path.join(tempfile.gettempdir(), "subjects_export.xlsx")
        _build_subject_template(
            path,
            {
                "College Name": context["college_name"],
                "Branch": context["branch"],
                "Semester": context["semester"],
                "Curriculum": context["regulation"],
                "Curid": context["curid"],
                "Academic Year": context["academic_year"],
            },
        )

        wb = load_workbook(path)
        ws = wb.active
        header_row = len(SUBJECT_TEMPLATE_METADATA_ROWS) + 1
        start_row = header_row + 1

        rows = Subject.objects.filter(
            academic_year=context["academic_year"],
            branch=context["branch"],
            semester=context["semester"],
            regulation=context["regulation"],
        ).order_by("subject_id")

        for i, s in enumerate(rows, start=1):
            r = start_row + i - 1
            ws.cell(row=r, column=1, value=i)  # SubOrder
            ws.cell(row=r, column=2, value=s.subject_id)  # SubCode
            ws.cell(row=r, column=6, value=s.subject_name)  # Name
            ws.cell(row=r, column=13, value=s.subject_type)  # Subtype
            ws.cell(row=r, column=14, value=float(s.credits))  # Credits
            ws.cell(row=r, column=15, value=s.subject_type)  # SubjectCategory
            for c in range(1, len(SUBJECT_EXCEL_COLUMNS) + 1):
                ws.cell(row=r, column=c).protection = Protection(locked=False)

        wb.save(path)
        wb.close()

        return FileResponse(open(path, "rb"), as_attachment=True, filename="subjects_export.xlsx")


class TimetableContextView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        academic_year = _clean(request.query_params.get("academic_year"))
        branch = _clean(request.query_params.get("branch"))
        semester = _clean(request.query_params.get("semester"))
        section = _clean(request.query_params.get("section")).upper()
        regulation = _clean(request.query_params.get("regulation")).upper()

        if not academic_year or not branch or not semester:
            return Response(
                {"success": False, "message": "academic_year, branch and semester are required"},
                status=400,
            )

        subject_qs = Subject.objects.filter(academic_year=academic_year, branch=branch, semester=semester)
        if regulation:
            subject_qs = subject_qs.filter(regulation=regulation)
        subjects = list(
            subject_qs.order_by("subject_id").values("id", "subject_id", "subject_name", "subject_type", "credits", "regulation")
        )
        for s in subjects:
            s["subject_code"] = s["subject_id"]
            s["credits"] = str(s["credits"])

        mapped_sections = list(
            TimetableMapping.objects.filter(
                academic_year=academic_year,
                branch=branch,
                semester=semester,
            )
            .values_list("section", flat=True)
            .distinct()
        )
        sections = sorted(set(DEFAULT_SECTIONS + mapped_sections))

        mappings = list(
            TimetableMapping.objects.filter(
                academic_year=academic_year,
                branch=branch,
                semester=semester,
            )
            .select_related("faculty", "subject")
            .order_by("week_day", "period_no")
        )
        mapping_rows = [
            {
                "id": m.id,
                "section": m.section,
                "week_day": m.week_day,
                "period_no": m.period_no,
                "period_label": m.period_label,
                "period_time": m.period_time,
                "subject_id": m.subject.subject_id,
                "subject_code": m.subject.subject_id,
                "subject_name": m.subject.subject_name,
                "subject_type": m.subject.subject_type,
                "credits": str(m.subject.credits),
                "faculty_id_no": m.faculty.id_no,
                "faculty_name": m.faculty.full_name,
                "faculty_email": m.faculty.email or "",
            }
            for m in mappings
        ]

        subject_faculty_qs = SubjectFacultyMapping.objects.filter(
            academic_year=academic_year,
            branch=branch,
            semester=semester,
        ).select_related("subject", "faculty")
        if regulation:
            subject_faculty_qs = subject_faculty_qs.filter(regulation=regulation)
        if section:
            subject_faculty_qs = subject_faculty_qs.filter(section=section)

        subject_faculty_rows = [
            {
                "slot_key": m.slot_key,
                "section": m.section,
                "subject_pk": m.subject_id,
                "subject_code": m.subject.subject_id,
                "subject_name": m.subject.subject_name,
                "faculty_id_no": m.faculty.id_no,
                "faculty_name": m.faculty.full_name or "",
            }
            for m in subject_faculty_qs.order_by("slot_key")
        ]

        saved_qs = Timetable.objects.filter(
            academic_year=academic_year,
            branch=branch,
            semester=semester,
        )
        if regulation:
            saved_qs = saved_qs.filter(regulation=regulation)
        if section:
            saved_qs = saved_qs.filter(section=section)
        saved_rows = _serialize_saved_timetable(saved_qs)

        return Response(
            {
                "success": True,
                "sections": sections,
                "subjects": subjects,
                "mappings": mapping_rows,
                "subject_faculty_mappings": subject_faculty_rows,
                "saved_timetable": saved_rows,
                "has_saved_timetable": len(saved_rows) > 0,
            },
            status=200,
        )


class TimetableFacultySearchView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        q = _clean(request.query_params.get("q"))
        today = timezone.now().date()
        qs = Faculty.objects.filter(Q(relieving_date__isnull=True) | Q(relieving_date__gt=today))

        if q:
            qs = qs.filter(Q(id_no__icontains=q) | Q(full_name__icontains=q) | Q(email__icontains=q))

        rows = []
        for f in qs.order_by("id_no")[:30]:
            rows.append(
                {
                    "id_no": f.id_no,
                    "full_name": f.full_name or "",
                    "email": f.email or "",
                    "department": f.department or "",
                    "designation": f.designation or "",
                }
            )
        return Response({"success": True, "count": len(rows), "results": rows}, status=200)


class TimetableMappingCreateView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        academic_year = _clean(request.data.get("academic_year"))
        branch = _clean(request.data.get("branch"))
        semester = _clean(request.data.get("semester"))
        regulation = _clean(request.data.get("regulation")).upper()
        section = _clean(request.data.get("section")).upper()
        week_day = _clean(request.data.get("week_day"))
        subject_pk = request.data.get("subject_id")
        faculty_id_no = _clean(request.data.get("faculty_id_no")).upper()
        period_no = int(request.data.get("period_no") or 0)

        if not academic_year or not branch or not semester:
            return Response({"success": False, "message": "academic_year, branch and semester are required"}, status=400)
        if not section:
            return Response({"success": False, "message": "section is required"}, status=400)
        if week_day not in VALID_DAYS:
            return Response({"success": False, "message": "week_day is invalid"}, status=400)
        if period_no not in PERIOD_MASTER:
            return Response({"success": False, "message": "period_no is invalid"}, status=400)
        if not subject_pk:
            return Response({"success": False, "message": "subject_id is required"}, status=400)
        if not faculty_id_no:
            return Response({"success": False, "message": "faculty_id_no is required"}, status=400)

        subject_qs = Subject.objects.filter(
            id=subject_pk,
            academic_year=academic_year,
            branch=branch,
            semester=semester,
        )
        if regulation:
            subject_qs = subject_qs.filter(regulation=regulation)
        subject = subject_qs.first()
        if not subject:
            return Response({"success": False, "message": "Selected subject not found in this context"}, status=404)

        today = timezone.now().date()
        faculty = Faculty.objects.filter(id_no=faculty_id_no).first()
        if not faculty:
            return Response({"success": False, "message": "Faculty not found"}, status=404)
        if faculty.relieving_date and faculty.relieving_date <= today:
            return Response({"success": False, "message": "Selected faculty is relieved/inactive"}, status=400)

        period_label, period_time = PERIOD_MASTER[period_no]

        class_slot_conflict = TimetableMapping.objects.filter(
            academic_year=academic_year,
            branch=branch,
            semester=semester,
            section=section,
            week_day=week_day,
            period_no=period_no,
        ).first()
        if class_slot_conflict:
            return Response(
                {
                    "success": False,
                    "message": (
                        "This class slot is already mapped to "
                        f"{class_slot_conflict.faculty.full_name or class_slot_conflict.faculty.id_no}"
                    ),
                },
                status=409,
            )

        faculty_slot_conflict = TimetableMapping.objects.filter(
            faculty=faculty,
            week_day=week_day,
            period_no=period_no,
        ).first()
        if faculty_slot_conflict:
            return Response(
                {
                    "success": False,
                    "message": (
                        "This faculty is already assigned at the same day and period "
                        f"({faculty_slot_conflict.branch} {faculty_slot_conflict.semester} "
                        f"{faculty_slot_conflict.section})"
                    ),
                },
                status=409,
            )

        mapping = TimetableMapping.objects.create(
            academic_year=academic_year,
            branch=branch,
            semester=semester,
            section=section,
            week_day=week_day,
            period_no=period_no,
            period_label=period_label,
            period_time=period_time,
            subject=subject,
            faculty=faculty,
        )
        return Response(
            {
                "success": True,
                "message": "Timetable mapping saved successfully",
                "data": {
                    "id": mapping.id,
                    "section": mapping.section,
                    "week_day": mapping.week_day,
                    "period_no": mapping.period_no,
                    "period_label": mapping.period_label,
                    "period_time": mapping.period_time,
                    "subject_code": mapping.subject.subject_id,
                    "subject_id": mapping.subject.subject_id,
                    "subject_name": mapping.subject.subject_name,
                    "subject_type": mapping.subject.subject_type,
                    "credits": str(mapping.subject.credits),
                    "faculty_id_no": mapping.faculty.id_no,
                    "faculty_name": mapping.faculty.full_name,
                },
            },
            status=201,
        )


class SubjectFacultyMappingSaveView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        academic_year = _clean(request.data.get("academic_year"))
        branch = _clean(request.data.get("branch"))
        semester = _clean(request.data.get("semester"))
        section = _clean(request.data.get("section")).upper()
        entries = request.data.get("entries") or []

        if not academic_year or not branch or not semester or not section:
            return Response(
                {"success": False, "message": "academic_year, branch, semester and section are required"},
                status=400,
            )
        if not isinstance(entries, list) or not entries:
            return Response({"success": False, "message": "entries are required"}, status=400)

        subject_qs = Subject.objects.filter(
            academic_year=academic_year,
            branch=branch,
            semester=semester,
        )
        subject_by_pk = {s.id: s for s in subject_qs}
        today = timezone.now().date()

        prepared_rows = []
        seen_slots = set()
        for idx, row in enumerate(entries, start=1):
            slot_key = _clean(row.get("slot_key"))
            subject_pk = row.get("subject_id")
            faculty_id_no = _clean(row.get("faculty_id_no")).upper()

            if not slot_key:
                return Response({"success": False, "message": f"slot_key is required for row {idx}"}, status=400)
            if slot_key in seen_slots:
                return Response({"success": False, "message": f"Duplicate slot_key: {slot_key}"}, status=400)
            seen_slots.add(slot_key)

            try:
                subject_pk_int = int(subject_pk)
            except (TypeError, ValueError):
                return Response({"success": False, "message": f"Invalid subject_id for slot {slot_key}"}, status=400)
            subject = subject_by_pk.get(subject_pk_int)
            if not subject:
                return Response(
                    {"success": False, "message": f"Subject not found in selected context for slot {slot_key}"},
                    status=400,
                )

            if not faculty_id_no:
                return Response({"success": False, "message": f"faculty_id_no is required for slot {slot_key}"}, status=400)
            faculty = Faculty.objects.filter(id_no=faculty_id_no).first()
            if not faculty:
                return Response({"success": False, "message": f"Faculty not found for slot {slot_key}"}, status=400)
            if faculty.relieving_date and faculty.relieving_date <= today:
                return Response({"success": False, "message": f"Faculty is relieved/inactive for slot {slot_key}"}, status=400)

            prepared_rows.append(
                SubjectFacultyMapping(
                    academic_year=academic_year,
                    branch=branch,
                    semester=semester,
                    section=section,
                    regulation=subject.regulation,
                    slot_key=slot_key,
                    subject=subject,
                    faculty=faculty,
                )
            )

        with transaction.atomic():
            SubjectFacultyMapping.objects.filter(
                academic_year=academic_year,
                branch=branch,
                semester=semester,
                section=section,
            ).delete()
            SubjectFacultyMapping.objects.bulk_create(prepared_rows)

        return Response(
            {
                "success": True,
                "message": "Subject-faculty mapping saved successfully",
                "saved_count": len(prepared_rows),
            },
            status=200,
        )


class TimetableAutoGenerateView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        academic_year = _clean(request.data.get("academic_year"))
        branch = _clean(request.data.get("branch"))
        semester = _clean(request.data.get("semester"))
        section = _clean(request.data.get("section")).upper()
        regulation = _clean(request.data.get("regulation") or "R20").upper()

        if not academic_year or not branch or not semester or not section:
            return Response(
                {
                    "success": False,
                    "message": "academic_year, branch, semester and section are required",
                },
                status=400,
            )

        class_id = f"{academic_year}:{branch}:{semester}:{section}:{regulation}"
        try:
            timetable = auto_generate_timetable(
                class_id=class_id,
                academic_year=academic_year,
                branch=branch,
                semester=semester,
                section=section,
                regulation=regulation,
            )
        except TimetableGenerationError as exc:
            return Response({"success": False, "message": str(exc)}, status=400)
        except Exception:
            return Response(
                {"success": False, "message": "Unexpected error while generating timetable"},
                status=500,
            )

        return Response(
            {
                "success": True,
                "message": "Timetable generated successfully",
                "class_id": class_id,
                "timetable": timetable,
            },
            status=200,
        )


class TimetableSaveView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        academic_year = _clean(request.data.get("academic_year"))
        branch = _clean(request.data.get("branch"))
        semester = _clean(request.data.get("semester"))
        section = _clean(request.data.get("section")).upper()
        regulation = _clean(request.data.get("regulation") or "R20").upper()
        overwrite = bool(request.data.get("overwrite"))
        entries = request.data.get("entries") or []

        if not academic_year or not branch or not semester or not section:
            return Response(
                {"success": False, "message": "academic_year, branch, semester and section are required"},
                status=400,
            )
        if not isinstance(entries, list) or not entries:
            return Response({"success": False, "message": "entries are required"}, status=400)

        existing_qs = Timetable.objects.filter(
            academic_year=academic_year,
            branch=branch,
            semester=semester,
            section=section,
            regulation=regulation,
        )
        has_existing = existing_qs.exists()
        if has_existing and not overwrite:
            return Response(
                {
                    "success": False,
                    "exists": True,
                    "message": "Timetable already exists for selected context. Do you want to update it?",
                },
                status=409,
            )

        subject_qs = Subject.objects.filter(
            academic_year=academic_year,
            branch=branch,
            semester=semester,
            regulation=regulation,
        )
        subject_by_code = {s.subject_id.upper(): s for s in subject_qs}
        slot_map = _ensure_time_slots()

        prepared = {}
        for row in entries:
            day_full = _to_full_day(row.get("day"))
            period_no = int(row.get("period_no") or 0)
            subject_code = _clean(row.get("subject_code")).upper()
            if day_full not in VALID_DAYS:
                return Response({"success": False, "message": f"Invalid day: {row.get('day')}"}, status=400)
            if period_no not in PERIOD_MASTER:
                return Response({"success": False, "message": f"Invalid period_no: {period_no}"}, status=400)
            if subject_code not in subject_by_code:
                return Response(
                    {"success": False, "message": f"Invalid subject_code for selected context: {subject_code}"},
                    status=400,
                )
            prepared[(day_full, period_no)] = subject_by_code[subject_code]

        expected_slots = len(VALID_DAYS) * len(PERIOD_MASTER)
        if len(prepared) != expected_slots:
            return Response(
                {"success": False, "message": f"All {expected_slots} timetable slots must be filled before save"},
                status=400,
            )

        practical_ok, practical_msg = _validate_practical_continuity_for_timetable_entries(prepared)
        if not practical_ok:
            return Response({"success": False, "message": practical_msg}, status=409)

        ok, validation_msg = _validate_faculty_conflicts_for_timetable_entries(
            academic_year=academic_year,
            branch=branch,
            semester=semester,
            section=section,
            regulation=regulation,
            prepared=prepared,
        )
        if not ok:
            return Response({"success": False, "message": validation_msg}, status=409)

        with transaction.atomic():
            existing_qs.delete()
            rows = []
            for (day, period_no), subject in prepared.items():
                rows.append(
                    Timetable(
                        academic_year=academic_year,
                        branch=branch,
                        semester=semester,
                        section=section,
                        regulation=regulation,
                        day=day,
                        period_no=period_no,
                        subject=subject,
                        timeslot=slot_map[(day, period_no)],
                    )
                )
            Timetable.objects.bulk_create(rows)

        return Response(
            {
                "success": True,
                "message": "Timetable updated successfully" if has_existing else "Timetable saved successfully",
                "saved_count": len(prepared),
            },
            status=200,
        )


def _mid_context(request):
    return {
        "college_name": _clean(request.data.get("college_name") or request.query_params.get("college_name") or "SVR Engineering College"),
        "academic_year": _clean(request.data.get("academic_year") or request.query_params.get("academic_year")),
        "branch": _clean(request.data.get("branch") or request.query_params.get("branch")),
        "semester": _clean(request.data.get("semester") or request.query_params.get("semester")),
        "section": _clean(request.data.get("section") or request.query_params.get("section")).upper(),
        "exam_type": _clean(request.data.get("exam_type") or request.query_params.get("exam_type")).upper(),
        "regulation": _clean(request.data.get("regulation") or request.query_params.get("regulation") or "R20").upper(),
    }


def _mid_missing(context):
    missing = []
    for field in ["academic_year", "branch", "semester", "section", "exam_type", "regulation"]:
        if not context[field]:
            missing.append(field)
    return missing


def _build_mid_template(path, metadata, theory_subjects, students, marks_map=None):
    marks_map = marks_map or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "MidMarks"

    label_fill = PatternFill(fill_type="solid", start_color="EAF2FF", end_color="EAF2FF")
    header_fill = PatternFill(fill_type="solid", start_color="DDE7F3", end_color="DDE7F3")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for idx, key in enumerate(MID_TEMPLATE_METADATA_ROWS, start=1):
        ws.cell(row=idx, column=1, value=key).font = bold
        ws.cell(row=idx, column=1).fill = label_fill
        ws.cell(row=idx, column=2, value=metadata.get(key, ""))
        ws.cell(row=idx, column=1).protection = Protection(locked=True)
        ws.cell(row=idx, column=2).protection = Protection(locked=True)

    header_row = len(MID_TEMPLATE_METADATA_ROWS) + 1
    subject_headers = [f"{s.subject_id} - {s.subject_name}" for s in theory_subjects]
    columns = ["S.No", "Roll No"] + subject_headers
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.protection = Protection(locked=True)

    for i, student in enumerate(students, start=1):
        row_no = header_row + i
        ws.cell(row=row_no, column=1, value=i).protection = Protection(locked=True)
        ws.cell(row=row_no, column=2, value=student.hall_ticket_no).protection = Protection(locked=True)
        for s_idx, subject in enumerate(theory_subjects, start=3):
            mark_val = marks_map.get((student.hall_ticket_no, subject.subject_id))
            ws.cell(row=row_no, column=s_idx, value=mark_val if mark_val is not None else "").protection = Protection(locked=False)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    for idx in range(3, 3 + len(theory_subjects)):
        col = chr(ord("A") + idx - 1)
        ws.column_dimensions[col].width = 22

    ws.freeze_panes = f"A{header_row + 1}"
    ws.protection.sheet = True
    ws.protection.enable()
    wb.save(path)
    wb.close()


class MidMarksContextOptionsView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        now_year = datetime.now().year
        auto_years = [f"{y}-{y + 1}" for y in range(now_year - 2, now_year + 5)]
        db_years = list(Student.objects.exclude(academic_year="").values_list("academic_year", flat=True).distinct())
        years = sorted(set(auto_years + db_years))
        branches = sorted(
            set(
                list(Student.objects.exclude(branch="").values_list("branch", flat=True).distinct())
                + ["CSE", "ECE", "EEE", "MECH", "CIVIL", "CSE (AI)", "IT"]
            )
        )
        semesters = ["1-1", "1-2", "2-1", "2-2", "3-1", "3-2", "4-1", "4-2"]
        sections = sorted(set(DEFAULT_SECTIONS + list(Student.objects.exclude(section="").values_list("section", flat=True).distinct())))
        regulations = sorted(
            set(
                DEFAULT_REGULATIONS
                + list(
                    Subject.objects.exclude(regulation="")
                    .values_list("regulation", flat=True)
                    .distinct()
                )
            )
        )
        return Response(
            {
                "success": True,
                "academic_years": years,
                "branches": branches,
                "semesters": semesters,
                "sections": sections,
                "exam_types": DEFAULT_EXAM_TYPES,
                "regulations": regulations,
            },
            status=200,
        )


class MidMarksTemplateDownloadView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        context = _mid_context(request)
        missing = _mid_missing(context)
        if missing:
            return Response({"success": False, "message": "Required query params: " + ", ".join(missing)}, status=400)

        theory_subjects = list(
            Subject.objects.filter(
                academic_year=context["academic_year"],
                branch=context["branch"],
                semester=context["semester"],
                regulation=context["regulation"],
                subject_type="THEORY",
            ).order_by("subject_id")
        )
        if not theory_subjects:
            return Response({"success": False, "message": "No THEORY subjects found for selected context"}, status=400)

        students = list(
            Student.objects.filter(
                academic_year=context["academic_year"],
                branch=context["branch"],
                semester=context["semester"],
                section=context["section"],
            ).order_by("hall_ticket_no")
        )
        if not students:
            return Response({"success": False, "message": "No students found for selected context"}, status=400)

        path = os.path.join(tempfile.gettempdir(), "mid_marks_template.xlsx")
        _build_mid_template(
            path,
            {
                "College Name": context["college_name"],
                "Branch": context["branch"],
                "Section": context["section"],
                "Semester": context["semester"],
                "Academic Year": context["academic_year"],
            },
            theory_subjects,
            students,
        )
        return FileResponse(open(path, "rb"), as_attachment=True, filename="mid_marks_template.xlsx")


class MidMarksUploadView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        context = _mid_context(request)
        missing = _mid_missing(context)
        if missing:
            return Response({"success": False, "message": "Required fields: " + ", ".join(missing)}, status=400)

        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"success": False, "message": "Excel file is required"}, status=400)

        file_obj.seek(0)
        wb = load_workbook(file_obj, data_only=True)
        ws = wb.active
        metadata = {
            "college_name": _safe_text(ws.cell(row=1, column=2).value),
            "branch": _safe_text(ws.cell(row=2, column=2).value),
            "section": _safe_text(ws.cell(row=3, column=2).value).upper(),
            "semester": _safe_text(ws.cell(row=4, column=2).value),
            "academic_year": _safe_text(ws.cell(row=5, column=2).value),
        }
        wb.close()

        mismatches = []
        if metadata["branch"].lower() != context["branch"].lower():
            mismatches.append("branch")
        if metadata["section"].lower() != context["section"].lower():
            mismatches.append("section")
        if metadata["semester"].lower() != context["semester"].lower():
            mismatches.append("semester")
        if metadata["academic_year"].lower() != context["academic_year"].lower():
            mismatches.append("academic_year")
        if mismatches:
            return Response({"success": False, "message": "Template metadata mismatch: " + ", ".join(mismatches)}, status=400)

        file_obj.seek(0)
        df = pd.read_excel(file_obj, header=len(MID_TEMPLATE_METADATA_ROWS), dtype=str).fillna("")
        columns = list(df.columns)
        if len(columns) < 3 or columns[0] != "S.No" or columns[1] != "Roll No":
            return Response({"success": False, "message": "Invalid template header. Download a fresh template."}, status=400)

        theory_subjects = list(
            Subject.objects.filter(
                academic_year=context["academic_year"],
                branch=context["branch"],
                semester=context["semester"],
                regulation=context["regulation"],
                subject_type="THEORY",
            ).order_by("subject_id")
        )
        subject_by_header = {f"{s.subject_id} - {s.subject_name}": s for s in theory_subjects}
        missing_cols = [h for h in columns[2:] if h not in subject_by_header]
        if missing_cols:
            return Response({"success": False, "message": "Unknown subject columns in template", "columns": missing_cols[:10]}, status=400)

        created = 0
        updated = 0
        errors = []
        with transaction.atomic():
            for idx, row in df.iterrows():
                excel_row = int(idx) + len(MID_TEMPLATE_METADATA_ROWS) + 2
                roll_no = _clean(row.get("Roll No"))
                if not roll_no:
                    continue
                student_exists = Student.objects.filter(
                    hall_ticket_no=roll_no,
                    academic_year=context["academic_year"],
                    branch=context["branch"],
                    semester=context["semester"],
                    section=context["section"],
                ).exists()
                if not student_exists:
                    errors.append({"row": excel_row, "message": f"Student not found in selected context: {roll_no}"})
                    continue

                for col in columns[2:]:
                    subject = subject_by_header.get(col)
                    if not subject:
                        continue
                    raw = _clean(row.get(col))
                    if raw == "":
                        continue
                    try:
                        marks = Decimal(raw)
                    except (InvalidOperation, ValueError):
                        errors.append({"row": excel_row, "message": f"Invalid marks for {col}"})
                        continue

                    obj, is_created = MidMark.objects.update_or_create(
                        academic_year=context["academic_year"],
                        branch=context["branch"],
                        semester=context["semester"],
                        section=context["section"],
                        exam_type=context["exam_type"],
                        student_roll_no=roll_no,
                        subject=subject,
                        defaults={"marks": marks},
                    )
                    if is_created:
                        created += 1
                    else:
                        updated += 1

        return Response(
            {
                "success": len(errors) == 0,
                "message": "Mid marks uploaded successfully" if not errors else "Mid marks uploaded with row errors",
                "summary": {"created": created, "updated": updated, "errors": len(errors)},
                "errors": errors[:25],
            },
            status=200,
        )


class MidMarksExportView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        context = _mid_context(request)
        missing = _mid_missing(context)
        if missing:
            return Response({"success": False, "message": "Required query params: " + ", ".join(missing)}, status=400)

        theory_subjects = list(
            Subject.objects.filter(
                academic_year=context["academic_year"],
                branch=context["branch"],
                semester=context["semester"],
                regulation=context["regulation"],
                subject_type="THEORY",
            ).order_by("subject_id")
        )
        students = list(
            Student.objects.filter(
                academic_year=context["academic_year"],
                branch=context["branch"],
                semester=context["semester"],
                section=context["section"],
            ).order_by("hall_ticket_no")
        )
        if not theory_subjects or not students:
            return Response({"success": False, "message": "No subjects or students found for selected context"}, status=400)

        marks_qs = MidMark.objects.filter(
            academic_year=context["academic_year"],
            branch=context["branch"],
            semester=context["semester"],
            section=context["section"],
            exam_type=context["exam_type"],
        ).select_related("subject")
        marks_map = {(m.student_roll_no, m.subject.subject_id): float(m.marks) for m in marks_qs}

        path = os.path.join(tempfile.gettempdir(), "mid_marks_export.xlsx")
        _build_mid_template(
            path,
            {
                "College Name": context["college_name"],
                "Branch": context["branch"],
                "Section": context["section"],
                "Semester": context["semester"],
                "Academic Year": context["academic_year"],
            },
            theory_subjects,
            students,
            marks_map=marks_map,
        )
        return FileResponse(open(path, "rb"), as_attachment=True, filename="mid_marks_export.xlsx")
